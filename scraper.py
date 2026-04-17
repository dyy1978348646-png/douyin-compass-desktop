"""
抖音罗盘数据抓取器 - 浏览器自动化核心逻辑

完整流程：
  1. 加载 Cookie → 打开罗盘首页
  2. 等待用户完成所有前置操作（入口选择、登录、账号选择）
     程序不强行检测每一步，只等最终结果：进入仪表盘
  3. 进入仪表盘后保存 Cookie
  4. 根据入口类型（达人/店铺）导航到对应的数据模块
  5. 导出数据并保存

入口类型区别：
  - 达人入口：直播 → 直播复盘
  - 店铺入口：直播 → 实时直播数据
"""

import json
import logging
import os
import platform
import re
import shutil
import socket
import subprocess
import sys
import time
from datetime import date, datetime
from pathlib import Path
from urllib.parse import urlsplit

import pandas as pd
from playwright._impl._driver import compute_driver_executable, get_driver_env
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

from config_manager import COOKIE_FILE, DOWNLOAD_DIR, APP_DIR, PLAYWRIGHT_BROWSERS_DIR
from scraper_logic import (
    ACCOUNT_SELECTION_KEYWORDS,
    DASHBOARD_NAV_TEXTS,
    DATE_MODE_LAST_1_DAY,
    DATE_MODE_LAST_7_DAYS,
    PORTAL_SELECTION_KEYWORDS,
    SCENE_DISPLAY_NAMES,
    SCENE_HOME_OVERVIEW,
    SCENE_LIVE_REVIEW,
    SCENE_SHOP_LIVE_DATA,
    SCENE_UNKNOWN,
    SCENE_VIDEO_REVIEW,
    SCENE_VISIBLE_KEYWORDS,
    TASK_STATUS_CANCELLED,
    TASK_STATUS_EXPORTING,
    TASK_STATUS_FAILED,
    TASK_STATUS_PENDING,
    TASK_STATUS_PRECHECKING,
    TASK_STATUS_SELECTING_DATE,
    TASK_STATUS_SUCCESS,
    detect_scene_snapshot,
    detect_page_status,
    build_video_review_export_rows,
    choose_account_name,
    extract_metric_fragment,
    is_account_selection_page_snapshot,
    is_dashboard_page_snapshot,
    is_target_date_range_visible,
    is_video_review_detail_page_snapshot,
    normalize_export_row_date,
    parse_metric_value,
    resolve_requested_scene,
    resolve_target_date_range,
)
from scraper_storage import persist_exported_dataframe
from runtime_env import (
    build_hidden_subprocess_kwargs,
    configure_playwright_browser_env,
    has_browser_install,
)
from scraper_waits import wait_for_page_ready, wait_until

logger = logging.getLogger("douyin_rpa")


def ensure_playwright_browser_installed():
    """
    优先使用打包阶段预装的 Chromium，缺失时再回退到用户目录安装。
    """
    browser_root = configure_playwright_browser_env(PLAYWRIGHT_BROWSERS_DIR)

    if has_browser_install(browser_root):
        installed = next(path for path in browser_root.glob("chromium-*") if path.is_dir())
        logger.info(f"检测到本地 Chromium: {installed}")
        return

    if getattr(sys, "frozen", False):
        logger.warning("未检测到随应用打包的 Chromium，回退到用户目录执行安装")
    else:
        logger.info("开发环境未检测到 Chromium，准备执行本地安装...")

    browser_root.mkdir(parents=True, exist_ok=True)

    driver_executable, driver_cli = compute_driver_executable()
    env = get_driver_env()
    env["PLAYWRIGHT_BROWSERS_PATH"] = str(browser_root)

    completed = subprocess.run(
        [driver_executable, driver_cli, "install", "chromium"],
        env=env,
        capture_output=True,
        text=True,
        **build_hidden_subprocess_kwargs(),
    )

    if completed.returncode != 0:
        stderr = (completed.stderr or "").strip()
        stdout = (completed.stdout or "").strip()
        detail = stderr or stdout or "未知错误"
        raise RuntimeError(
            "Chromium 浏览器组件安装失败。"
            "请确认当前电脑可以联网，或稍后重试。\n"
            f"安装输出: {detail}"
        )

    logger.info("Chromium 浏览器组件安装完成")


def chrome_executable_candidates(
    system_name: str | None = None,
    home: Path | None = None,
    env: dict | None = None,
) -> list[Path]:
    """返回系统 Chrome 可执行文件候选路径。"""
    system_name = system_name or platform.system()
    home = home or Path.home()
    env = env or os.environ

    if system_name == "Darwin":
        return [
            Path("/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"),
            home / "Applications" / "Google Chrome.app" / "Contents" / "MacOS" / "Google Chrome",
        ]

    if system_name == "Windows":
        candidates: list[Path] = []
        for key in ("PROGRAMFILES", "PROGRAMFILES(X86)", "LOCALAPPDATA"):
            base = env.get(key)
            if not base:
                continue
            candidates.append(Path(base) / "Google" / "Chrome" / "Application" / "chrome.exe")
        return candidates

    return [
        Path("/usr/bin/google-chrome"),
        Path("/usr/bin/chromium-browser"),
        Path("/usr/bin/chromium"),
    ]


def chrome_user_data_candidates(
    system_name: str | None = None,
    home: Path | None = None,
    env: dict | None = None,
) -> list[Path]:
    """返回 Chrome 用户数据目录候选路径。"""
    system_name = system_name or platform.system()
    home = home or Path.home()
    env = env or os.environ

    if system_name == "Darwin":
        return [home / "Library" / "Application Support" / "Google" / "Chrome"]

    if system_name == "Windows":
        local_app_data = env.get("LOCALAPPDATA")
        if local_app_data:
            return [Path(local_app_data) / "Google" / "Chrome" / "User Data"]
        return [home / "AppData" / "Local" / "Google" / "Chrome" / "User Data"]

    return [
        home / ".config" / "google-chrome",
        home / ".config" / "chromium",
    ]


def first_existing_path(candidates: list[Path]) -> Path | None:
    """返回第一个存在的路径。"""
    return next((path for path in candidates if path.exists()), None)

# ============================================================
# Cookie 管理
# ============================================================

# ============================================================
# 核心抓取类
# ============================================================
class DouyinCompassScraper:
    """封装抖音罗盘的浏览器自动化操作。"""

    # 入口类型常量
    PORTAL_CREATOR = "creator"   # 达人入口
    PORTAL_SHOP = "shop"         # 店铺入口

    def __init__(self, config: dict):
        self.config = config
        self.pw = None
        self.browser = None
        self.context = None
        self.page = None
        self.chrome_process = None
        self._cancelled = False
        self.account_name = ""  # 当前登录的账号名称
        self.scene_id = ""
        self.target_date_range = None
        self.actual_date_range = None
        self.date_scope = None
        self.task_status = TASK_STATUS_PENDING
        self.task_id = datetime.now().strftime("%Y%m%d%H%M%S")

    def cancel(self):
        """外部调用以取消正在执行的任务。"""
        self._cancelled = True

    def _check_cancel(self):
        if self._cancelled:
            raise InterruptedError("任务已被用户取消")

    @property
    def portal_type(self) -> str:
        """获取入口类型配置，默认达人入口。"""
        return self.config.get("portal_type", self.PORTAL_CREATOR)

    def run_switch_account(self) -> dict:
        """
        切换账号流程（手动确认模式）：
          1. 打开浏览器，导航到罗盘页面
          2. 等待用户在浏览器中完成账号切换
          3. 用户在 GUI 中点击「确认已切换」后关闭浏览器
        不做任何自动检测，完全由用户控制节奏。
        """
        self._cancelled = False
        self._switch_confirmed = False
        result = {"success": False, "message": ""}

        with sync_playwright() as pw:
            self.pw = pw
            try:
                ensure_playwright_browser_installed()
                self._start_browser()
                self._check_cancel()

                # 关闭旧标签页，保持干净
                self._close_extra_tabs()

                url = self.config.get("compass_url", "https://compass.jinritemai.com")
                logger.info(f"打开罗盘: {url}")
                self.page.goto(url, wait_until="networkidle")
                wait_for_page_ready(self.page)

                logger.info("=" * 50)
                logger.info("请在浏览器中完成账号切换：")
                logger.info("  1. 等待罗盘页面加载完成")
                logger.info("  2. 点击右上角头像 → 切换账号")
                logger.info("  3. 选择要使用的新账号")
                logger.info("  4. 等待新账号的罗盘页面加载完成")
                logger.info("  5. 回到本程序，点击「确认已切换」按钮")
                logger.info("=" * 50)

                # 等待用户在 GUI 中点击「确认已切换」
                timeout = self.config.get("login_wait_timeout", 300)
                def switch_confirmed() -> bool:
                    self._check_cancel()
                    return self._switch_confirmed

                try:
                    wait_until(
                        switch_confirmed,
                        timeout_seconds=timeout,
                        interval_seconds=1,
                        timeout_message=f"等待超时（{timeout}秒）",
                    )
                    logger.info("用户确认账号切换完成!")
                    result["success"] = True
                    result["message"] = "账号切换成功"
                except TimeoutError:
                    result["message"] = f"等待超时（{timeout}秒）"
                    logger.warning(result["message"])

            except InterruptedError:
                result["message"] = "已取消"
                logger.info("账号切换已取消")
            except Exception as e:
                result["message"] = f"失败: {e}"
                logger.exception(f"账号切换失败: {e}")
            finally:
                self._close()

        return result

    def confirm_switch(self):
        """GUI 调用：用户确认账号已切换完成。"""
        self._switch_confirmed = True

    def run_precheck(self) -> dict:
        """仅执行预检查：登录状态、目标场景进入和页面识别，不导出数据。"""
        self._cancelled = False
        self.task_id = datetime.now().strftime("%Y%m%d%H%M%S")
        self.task_status = TASK_STATUS_PENDING
        self.scene_id = resolve_requested_scene(self.config)
        self.target_date_range = None
        self.actual_date_range = None
        self.date_scope = None

        result = {
            "success": False,
            "message": "",
            "task_id": self.task_id,
            "task_status": self.task_status,
            "scene_id": self.scene_id,
            "scene_name": SCENE_DISPLAY_NAMES.get(self.scene_id, self.scene_id),
            "target_date_range": None,
            "detected_scene": None,
            "account_name": "",
        }

        with sync_playwright() as pw:
            self.pw = pw
            try:
                self.target_date_range = resolve_target_date_range(self.config)
                result["target_date_range"] = self.target_date_range.to_payload()
                ensure_playwright_browser_installed()
                self._start_browser()
                self._check_cancel()
                self.task_status = TASK_STATUS_PRECHECKING
                self._open_and_wait_for_dashboard()
                self._check_cancel()
                self.account_name = self._get_account_name()
                result["account_name"] = self.account_name
                self._navigate_to_target_scene()
                self._check_cancel()
                detection = self._assert_scene_ready(self.scene_id)
                self.task_status = TASK_STATUS_SUCCESS
                result["task_status"] = self.task_status
                result["detected_scene"] = detection.to_payload()
                result["success"] = True
                result["message"] = "页面检查通过，可以开始执行抓取。"
            except InterruptedError:
                self.task_status = TASK_STATUS_CANCELLED
                result["task_status"] = self.task_status
                result["message"] = "页面检查已取消"
            except Exception as e:
                self.task_status = TASK_STATUS_FAILED
                result["task_status"] = self.task_status
                result["message"] = f"页面检查失败: {e}"
                logger.exception("页面预检查失败: %s", e)
            finally:
                self._close()

        return result

    def _close_extra_tabs(self):
        """
        关闭多余的标签页，只保留一个。
        持久化浏览器启动时会恢复上次所有标签页，
        旧标签页上的仪表盘会干扰切换账号的检测逻辑。
        """
        pages = self.context.pages
        if len(pages) <= 1:
            return

        logger.info(f"检测到 {len(pages)} 个标签页，关闭多余标签页...")
        # 保留第一个页面，关闭其余的
        for page in pages[1:]:
            try:
                page.close()
            except Exception:
                pass

        # 确保 self.page 指向唯一剩余的页面
        if self.context.pages:
            self.page = self.context.pages[0]
        else:
            self.page = self.context.new_page()

        logger.info("已清理旧标签页")

    def run(self) -> dict:
        """执行完整抓取流程，返回结果摘要。"""
        self._cancelled = False
        self.task_id = datetime.now().strftime("%Y%m%d%H%M%S")
        self.task_status = TASK_STATUS_PENDING
        self.scene_id = resolve_requested_scene(self.config)
        self.target_date_range = None
        self.actual_date_range = None
        self.date_scope = None
        result = {
            "success": False,
            "rows": 0,
            "message": "",
            "filepath": "",
            "csv_path": "",
            "task_id": self.task_id,
            "task_status": self.task_status,
            "scene_id": self.scene_id,
            "scene_name": SCENE_DISPLAY_NAMES.get(self.scene_id, self.scene_id),
            "target_date_range": None,
            "actual_date_range": None,
        }

        portal_name = "达人" if self.portal_type == self.PORTAL_CREATOR else "店铺"
        logger.info(f"当前入口类型: {portal_name}")

        with sync_playwright() as pw:
            self.pw = pw
            try:
                self.target_date_range = resolve_target_date_range(self.config)
                result["target_date_range"] = self.target_date_range.to_payload()
                logger.info(
                    "本次任务: 场景=%s, 日期=%s (%s ~ %s)",
                    SCENE_DISPLAY_NAMES.get(self.scene_id, self.scene_id),
                    self.target_date_range.label,
                    self.target_date_range.start.isoformat(),
                    self.target_date_range.end.isoformat(),
                )
                ensure_playwright_browser_installed()
                self._start_browser()
                self._check_cancel()
                self.task_status = TASK_STATUS_PRECHECKING
                self._open_and_wait_for_dashboard()
                self._check_cancel()
                self.account_name = self._get_account_name()

                # ============================================================
                # CREATOR 达人入口：统一抓取流程（首页 + 直播复盘 + 短视频）
                # 重要：不走 500-507 行的通用导航，因为 CREATOR 的每个子模块
                # 都有自己独立的导航+日期选择，走统一流程更稳定
                # ============================================================
                if self.portal_type == self.PORTAL_CREATOR:
                    import openpyxl
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    from scraper_storage import _safe_filename_part

                    self.task_status = TASK_STATUS_EXPORTING
                    self.actual_date_range = self.target_date_range
                    target_date = self.target_date_range
                    date_str = target_date.start.strftime("%Y-%m-%d") if target_date else datetime.now().strftime("%Y-%m-%d")
                    account_safe = _safe_filename_part(self.account_name or "未知账号", "未知账号")
                    xlsx_name = f"罗盘数据抓取_{date_str}_{account_safe}.xlsx"
                    xlsx_path = DOWNLOAD_DIR / xlsx_name

                    sheets_data = []

                    # 1. 直播整体（首页 → 整体概括）
                    try:
                        logger.info("--- 开始抓取 [直播整体]（首页整体概括）---")
                        self.scene_id = SCENE_HOME_OVERVIEW
                        self._navigate_home_overview()
                        wait_for_page_ready(self.page, timeout_ms=8000)
                        self._apply_date_selection()
                        self._check_cancel()
                        self._assert_date_selection_applied(self.target_date_range)
                        summary_metrics = self._extract_home_overview_summary_metrics()
                        df_overall = self._build_live_overall_summary(summary_metrics)
                        logger.info("[直播整体] 抓取完成: %s 行", len(df_overall))
                        sheets_data.append({"name": "直播整体", "df": df_overall})
                    except Exception as exc:
                        logger.warning("[直播整体] 抓取失败: %s", exc)
                        sheets_data.append({"name": "直播整体", "df": None})

                    self._check_cancel()

                    # 2. 渠道明细（直播复盘页面）
                    try:
                        logger.info("--- 开始抓取 [渠道明细]（直播复盘页面）---")
                        self.scene_id = SCENE_LIVE_REVIEW
                        self._navigate_creator_live_review()
                        wait_for_page_ready(self.page, timeout_ms=8000)
                        self._apply_date_selection()
                        self._check_cancel()
                        self._assert_date_selection_applied(self.target_date_range)
                        df_channel, _ = self._build_live_review_export_data()
                        logger.info("[渠道明细] 抓取完成: %s 行", len(df_channel))
                        sheets_data.append({"name": "渠道明细", "df": df_channel})

                        # 2.1 抓取渠道分析漏斗表格（曝光/观看/点击率）
                        self._check_cancel()
                        try:
                            df_channel_analysis = self._extract_live_review_channel_analysis_as_df()
                            if not df_channel_analysis.empty:
                                crawl_time = datetime.now().isoformat()
                                if "账号名称" not in df_channel_analysis.columns:
                                    df_channel_analysis.insert(0, "账号名称", self.account_name or "")
                                if "抓取时间" not in df_channel_analysis.columns:
                                    df_channel_analysis.insert(1, "抓取时间", crawl_time)
                                if "目标日期" not in df_channel_analysis.columns:
                                    df_channel_analysis.insert(2, "目标日期", (
                                        self.target_date_range.start.isoformat()
                                        if self.target_date_range else ""
                                    ))
                                logger.info("[渠道分析] 抓取完成: %s 行", len(df_channel_analysis))
                                sheets_data.append({"name": "渠道分析", "df": df_channel_analysis})
                            else:
                                logger.warning("[渠道分析] 未获取到数据")
                        except Exception as exc:
                            logger.warning("[渠道分析] 抓取失败: %s", exc)

                    except Exception as exc:
                        logger.warning("[渠道明细] 抓取失败: %s", exc)
                        sheets_data.append({"name": "渠道明细", "df": None})

                    self._check_cancel()

                    # 3. 短视频引流（视频复盘页面）
                    try:
                        logger.info("--- 开始抓取 [短视频引流] ---")
                        self.scene_id = SCENE_VIDEO_REVIEW
                        self._navigate_video_review()
                        wait_for_page_ready(self.page, timeout_ms=8000)
                        self._scroll_to_top()
                        self.page.wait_for_timeout(2000)
                        self._check_cancel()
                        self._apply_video_review_date_input()
                        wait_for_page_ready(self.page, timeout_ms=5000)
                        df_video, _ = self._export_video_review_metrics()
                        logger.info("[短视频引流] 抓取完成: %s 行", len(df_video))
                        sheets_data.append({"name": "短视频引流", "df": df_video})
                    except Exception as exc:
                        logger.warning("[短视频引流] 抓取失败: %s", exc)
                        sheets_data.append({"name": "短视频引流", "df": None})

                    self.scene_id = SCENE_SHOP_LIVE_DATA  # 恢复原值
                    self._write_excel_unified(xlsx_path, sheets_data)
                    logger.info("Excel 已保存: %s", xlsx_path)

                    # 合并 DataFrame（防御：不同 sheet 列名可能重复/不同）
                    valid_dfs = [s["df"] for s in sheets_data if s["df"] is not None and not s["df"].empty]

                    if valid_dfs:
                        # 如果只有一个，直接用；多个则 concat
                        if len(valid_dfs) == 1:
                            df = valid_dfs[0].copy()
                        else:
                            # 清理重复列名：同名列只保留第一个
                            cleaned = []
                            for d in valid_dfs:
                                if d.columns.nunique() != len(d.columns):
                                    d = d.loc[:, ~d.columns.duplicated()]
                                cleaned.append(d)
                            if not cleaned:
                                df = pd.DataFrame()
                            else:
                                df = pd.concat(cleaned, ignore_index=True, join="outer")
                    else:
                        df = pd.DataFrame()
                    filepath = xlsx_path
                    csv_path = ""
                    result["xlsx_path"] = str(xlsx_path)
                    total_rows = sum(len(s["df"]) for s in sheets_data if s["df"] is not None and not s["df"].empty)
                    result["message"] = f"成功抓取 {total_rows} 行数据（直播整体 + 渠道明细 + 短视频引流）"
                else:
                    # 非 CREATOR 入口：走通用流程（原有逻辑）
                    self._navigate_to_target_scene()
                    self._check_cancel()
                    self._assert_scene_ready(self.scene_id)
                    self._assert_video_review_detail_page_context()
                    self.task_status = TASK_STATUS_SELECTING_DATE
                    self._apply_date_selection()
                    self._check_cancel()
                    self.actual_date_range = self._assert_date_selection_applied(self.target_date_range)
                    self.task_status = TASK_STATUS_EXPORTING
                    if self.scene_id == SCENE_HOME_OVERVIEW:
                        df, csv_path = self._export_home_overview_metrics()
                        filepath = None
                    elif self.scene_id == SCENE_VIDEO_REVIEW:
                        df, csv_path = self._export_video_review_metrics()
                        filepath = None
                    else:
                        filepath = self._export_data()
                        self._check_cancel()
                        df, csv_path = self._process_and_save(filepath)

                self.task_status = TASK_STATUS_SUCCESS
                result["success"] = True
                result["rows"] = len(df)
                result["filepath"] = str(filepath) if filepath else ""
                result["csv_path"] = str(csv_path)
                result["account_name"] = self.account_name
                result["task_status"] = self.task_status
                result["actual_date_range"] = self.actual_date_range.to_payload()
                result["message"] = f"成功抓取 {len(df)} 行数据"
                if self.account_name:
                    result["message"] += f"（账号: {self.account_name}）"
                logger.info(result["message"])

            except InterruptedError:
                self.task_status = TASK_STATUS_CANCELLED
                result["task_status"] = self.task_status
                result["message"] = "任务已取消"
                logger.info(result["message"])
            except Exception as e:
                self.task_status = TASK_STATUS_FAILED
                result["task_status"] = self.task_status
                if getattr(e, "csv_path", None):
                    result["csv_path"] = str(e.csv_path)
                result["message"] = f"失败: {e}"
                logger.exception(f"任务执行失败: {e}")
            finally:
                self._close()

        return result

    # ------ 内部方法 ------

    def _start_browser(self):
        """
        启动浏览器。
        """
        launch_mode = (self.config.get("browser_launch_mode") or "auto").strip().lower()

        if launch_mode != "persistent":
            try:
                self._start_browser_with_profile_clone()
                return
            except Exception as exc:
                if launch_mode == "clone_cdp":
                    raise
                logger.warning("真实 Chrome 副本启动失败，回退到持久化上下文: %s", exc)

        self._start_browser_with_persistent_context()

    def _start_browser_with_persistent_context(self):
        """回退方案：使用 Playwright 持久化上下文直接启动浏览器。"""
        headless = self.config.get("headless", False)
        configured_profile_dir = (self.config.get("browser_profile_dir") or "").strip()
        browser_channel = (self.config.get("browser_channel") or "").strip()

        # 浏览器配置文件目录（默认仍保存在程序目录下，避免直接污染真实浏览器 profile）
        user_data_dir = (
            Path(configured_profile_dir).expanduser()
            if configured_profile_dir
            else APP_DIR / "browser_profile"
        )
        user_data_dir.mkdir(parents=True, exist_ok=True)

        launch_options = {
            "user_data_dir": str(user_data_dir),
            "headless": headless,
            "accept_downloads": True,
            "viewport": {"width": 1920, "height": 1080},
            "locale": "zh-CN",
            "args": [
                "--disable-blink-features=AutomationControlled",
                "--disable-dev-shm-usage",
                "--no-first-run",
                "--no-default-browser-check",
            ],
        }
        if browser_channel:
            launch_options["channel"] = browser_channel

        logger.info(
            "启动浏览器 (headless=%s, channel=%s, 持久化配置=%s)",
            headless,
            browser_channel or "chromium",
            user_data_dir,
        )

        try:
            self.context = self.pw.chromium.launch_persistent_context(**launch_options)
        except Exception as exc:
            if browser_channel:
                logger.warning("系统 Chrome 启动失败，回退到 Playwright Chromium: %s", exc)
                launch_options.pop("channel", None)
                self.context = self.pw.chromium.launch_persistent_context(**launch_options)
            else:
                raise

        self.context.add_init_script(
            """
            Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
            window.chrome = window.chrome || { runtime: {} };
            """
        )
        self.context.set_default_timeout(self.config.get("browser_timeout", 60000))

        # persistent context 自带一个页面，如果没有就新建
        if self.context.pages:
            self.page = self.context.pages[0]
        else:
            self.page = self.context.new_page()

    def _resolve_system_chrome_executable(self) -> Path | None:
        configured = (self.config.get("chrome_executable_path") or "").strip()
        if configured:
            path = Path(configured).expanduser()
            return path if path.exists() else None
        return first_existing_path(chrome_executable_candidates())

    def _resolve_real_chrome_user_data_dir(self) -> Path | None:
        configured = (self.config.get("chrome_user_data_dir") or "").strip()
        if configured:
            path = Path(configured).expanduser()
            return path if path.exists() else None
        return first_existing_path(chrome_user_data_candidates())

    def _copy_path_if_exists(self, source: Path, target: Path):
        if not source.exists():
            return
        target.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source, target)

    def _copy_tree_best_effort(self, source: Path, target: Path):
        ignored_dirs = {"Cache", "Code Cache", "GPUCache", "GrShaderCache", "GraphiteDawnCache"}

        for root, dirs, files in os.walk(source):
            root_path = Path(root)
            relative = root_path.relative_to(source)
            if any(part in ignored_dirs for part in relative.parts):
                dirs[:] = []
                continue

            dirs[:] = [
                name for name in dirs
                if name not in ignored_dirs and not name.startswith("Singleton")
            ]

            target_dir = target / relative
            target_dir.mkdir(parents=True, exist_ok=True)

            for filename in files:
                if filename.startswith("Singleton") or filename.endswith(".lock"):
                    continue
                src_file = root_path / filename
                dst_file = target_dir / filename
                try:
                    shutil.copy2(src_file, dst_file)
                except OSError as exc:
                    logger.debug("跳过复制文件 %s: %s", src_file, exc)

    def _prepare_runtime_profile_clone(self, source_root: Path, profile_name: str) -> Path:
        runtime_root = APP_DIR / "browser_profile_runtime"
        if runtime_root.exists():
            shutil.rmtree(runtime_root, ignore_errors=True)
        runtime_root.mkdir(parents=True, exist_ok=True)

        self._copy_path_if_exists(source_root / "Local State", runtime_root / "Local State")
        self._copy_path_if_exists(source_root / "Last Version", runtime_root / "Last Version")

        source_profile = source_root / profile_name
        if not source_profile.exists():
            raise RuntimeError(f"未找到 Chrome 配置目录: {source_profile}")

        self._copy_tree_best_effort(source_profile, runtime_root / profile_name)
        return runtime_root

    def _pick_free_port(self) -> int:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            sock.bind(("127.0.0.1", 0))
            return int(sock.getsockname()[1])

    def _wait_for_cdp_port(self, port: int, timeout_seconds: int = 20):
        deadline = time.time() + timeout_seconds
        while time.time() < deadline:
            try:
                with socket.create_connection(("127.0.0.1", port), timeout=1):
                    return
            except OSError:
                time.sleep(0.5)
        raise RuntimeError(f"Chrome 调试端口未就绪: {port}")

    def _start_browser_with_profile_clone(self):
        """主方案：复制真实 Chrome profile，再通过 CDP 接管浏览器。"""
        chrome_executable = self._resolve_system_chrome_executable()
        if not chrome_executable:
            raise RuntimeError("未检测到系统 Chrome 可执行文件")

        source_root = self._resolve_real_chrome_user_data_dir()
        if not source_root:
            raise RuntimeError("未检测到系统 Chrome 用户数据目录")

        profile_name = (self.config.get("chrome_profile_name") or "Default").strip() or "Default"
        runtime_root = self._prepare_runtime_profile_clone(source_root, profile_name)
        port = self._pick_free_port()

        logger.info(
            "启动系统 Chrome（真实 profile 副本）: source=%s, profile=%s, runtime=%s",
            source_root,
            profile_name,
            runtime_root,
        )

        launch_cmd = [
            str(chrome_executable),
            f"--remote-debugging-port={port}",
            f"--user-data-dir={runtime_root}",
            f"--profile-directory={profile_name}",
            "--no-first-run",
            "--no-default-browser-check",
        ]

        self.chrome_process = subprocess.Popen(
            launch_cmd,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            **build_hidden_subprocess_kwargs(),
        )
        self._wait_for_cdp_port(port)

        self.browser = self.pw.chromium.connect_over_cdp(f"http://127.0.0.1:{port}")
        self.context = self.browser.contexts[0] if self.browser.contexts else self.browser.new_context()
        self.context.set_default_timeout(self.config.get("browser_timeout", 60000))

        if self.context.pages:
            self.page = self.context.pages[0]
        else:
            self.page = self.context.new_page()

    def _open_and_wait_for_dashboard(self):
        """
        打开罗盘网站，等待用户完成所有前置操作后进入仪表盘。

        不再逐步检测"当前在哪个页面"，而是：
          1. 打开网站
          2. 每隔几秒检查一次是否已到达仪表盘
          3. 期间持续在日志中提示当前状态
          4. 到达仪表盘后继续执行

        这样无论中间经过多少步（入口选择、登录、账号选择、验证码等），
        程序都不会误判，只要最终进入仪表盘就行。
        """
        url = self.config.get("compass_url", "https://compass.jinritemai.com")

        # 浏览器登录状态由当前启动策略决定：
        # - 真实 profile 副本模式：复用本机 Chrome 的已登录状态
        # - 持久化上下文模式：复用程序自己的 browser_profile

        logger.info(f"打开罗盘: {url}")
        self.page.goto(url, wait_until="networkidle")
        wait_for_page_ready(self.page)

        # 快速检查：若当前 profile 已登录，可能直接进入仪表盘
        if self._is_on_dashboard():
            logger.info("已登录，直接进入仪表盘")
            return

        # 需要用户手动操作
        logger.info("=" * 50)
        logger.info("请在浏览器中完成以下操作：")
        logger.info("  1. 选择入口（达人入口 / 店铺入口）")
        logger.info("  2. 手机验证码登录（如需要）")
        logger.info("  3. 选择账号（如有多个）")
        logger.info("完成后程序会自动继续，请勿关闭浏览器窗口")
        logger.info("若使用真实 Chrome 副本模式，请优先保持你本机 Chrome 已登录")
        logger.info("=" * 50)

        timeout = self.config.get("login_wait_timeout", 300)
        last_status = {"value": ""}

        def dashboard_ready() -> bool:
            self._check_cancel()

            # 检测当前状态并记录日志（避免重复打印）
            status = self._detect_page_status()
            if status != last_status["value"]:
                logger.info(f"当前页面状态: {status}")
                last_status["value"] = status

            return self._is_on_dashboard()

        wait_until(
            dashboard_ready,
            timeout_seconds=timeout,
            interval_seconds=1.5,
            timeout_message=(
                f"等待进入仪表盘超时（{timeout}秒）。"
                f"请确保在浏览器中完成登录和账号选择。"
            ),
        )
        logger.info("已进入仪表盘!")

    def _visible_keywords_for_page(self, page, keywords: tuple[str, ...]) -> list[str]:
        visible_keywords = []
        for keyword in keywords:
            try:
                element = page.query_selector(f'text="{keyword}"')
                if element and element.is_visible():
                    visible_keywords.append(keyword)
            except Exception:
                continue
        return visible_keywords

    def _collect_dashboard_visible_texts(self, page) -> list[str]:
        keywords = tuple(
            dict.fromkeys(
                PORTAL_SELECTION_KEYWORDS
                + ACCOUNT_SELECTION_KEYWORDS
                + DASHBOARD_NAV_TEXTS
                + SCENE_VISIBLE_KEYWORDS
            )
        )
        return self._visible_keywords_for_page(page, keywords)

    def _page_has_visible_text(self, texts: tuple[str, ...]) -> bool:
        selectors = []
        for text in texts:
            selectors.extend(
                [
                    f'text="{text}"',
                    f'a:has-text("{text}")',
                    f'button:has-text("{text}")',
                    f'span:has-text("{text}")',
                    f'div:has-text("{text}")',
                    f'li:has-text("{text}")',
                ]
            )

        for selector in selectors:
            try:
                element = self.page.query_selector(selector)
                if element and element.is_visible():
                    return True
            except Exception:
                continue
        return False

    def _expected_scene_marker_visible(self, expected_scene_id: str) -> bool:
        normalized_url = (self.page.url or "").lower()

        if expected_scene_id == SCENE_LIVE_REVIEW:
            return self._page_has_visible_text(("直播复盘",)) or (
                (("live" in normalized_url and "review" in normalized_url) or "replay" in normalized_url)
                and not self._page_has_visible_text(("视频复盘",))
            )

        if expected_scene_id == SCENE_VIDEO_REVIEW:
            visible_texts = self._visible_keywords_for_page(
                self.page,
                ("视频复盘", "视频明细", "视频榜单", "视频表现", "更多数据"),
            )
            return is_video_review_detail_page_snapshot(self.page.url, visible_texts)

        if expected_scene_id == SCENE_SHOP_LIVE_DATA:
            return self._page_has_visible_text(("实时直播数据",)) or "realtime" in normalized_url

        if expected_scene_id == SCENE_HOME_OVERVIEW:
            return (
                self._page_has_visible_text(("首页",))
                and self._page_has_visible_text(("整体概况", "整体概览"))
                and self._page_has_visible_text(("更多",))
            ) or any(keyword in normalized_url for keyword in ("overview", "dashboard", "/home"))

        return False

    def _detect_page_status(self) -> str:
        """检测当前页面大致状态，用于日志输出。"""
        visible_texts = self._collect_dashboard_visible_texts(self.page)
        return detect_page_status(self.page.url, visible_texts)

    def _is_on_dashboard(self) -> bool:
        """
        判断当前页面是否是罗盘仪表盘。
        会检查所有打开的标签页（登录过程中可能在新标签页打开仪表盘）。

        实际罗盘主页特征：
          - URL: compass.jinritemai.com/talent (达人端)
          - 顶部导航栏有: 首页、直播、短视频、图文、橱窗、交易、商品、商家、人群、市场
        """
        # 检查所有标签页，不只是当前页
        for page in self.context.pages:
            if self._check_page_is_dashboard(page):
                # 如果仪表盘在另一个标签页，切换过去
                if page != self.page:
                    logger.info("检测到仪表盘在新标签页，已切换")
                    self.page = page
                return True
        return False

    def _check_page_is_dashboard(self, page) -> bool:
        """检查指定页面是否是罗盘仪表盘。"""
        try:
            current_url = page.url
        except Exception:
            return False

        visible_texts = self._collect_dashboard_visible_texts(page)
        return is_dashboard_page_snapshot(current_url, visible_texts)

    def _detect_scene(self):
        visible_texts = self._collect_dashboard_visible_texts(self.page)
        return detect_scene_snapshot(self.page.url, visible_texts)

    def _assert_scene_ready(self, expected_scene_id: str):
        if self._expected_scene_marker_visible(expected_scene_id):
            detection = self._detect_scene()
            logger.info(
                "页面场景校验通过: %s (命中专用校验，通用识别=%s)",
                SCENE_DISPLAY_NAMES.get(expected_scene_id, expected_scene_id),
                SCENE_DISPLAY_NAMES.get(detection.scene_id, detection.scene_id),
            )
            return detection

        detection = self._detect_scene()
        if expected_scene_id == SCENE_HOME_OVERVIEW and detection.scene_id == SCENE_UNKNOWN:
            logger.warning(
                "首页场景识别不足，继续尝试抓取（将依赖首页模块定位结果）"
            )
            return detection
        if detection.scene_id != expected_scene_id:
            raise RuntimeError(
                "页面场景校验失败。"
                f"预期进入 {SCENE_DISPLAY_NAMES.get(expected_scene_id, expected_scene_id)}，"
                f"实际识别为 {SCENE_DISPLAY_NAMES.get(detection.scene_id, detection.scene_id)}。"
            )

        if detection.confidence == "low":
            raise RuntimeError(
                "当前页面场景识别置信度不足，为避免导出错误数据，本次任务已终止。"
            )

        logger.info(
            "页面场景校验通过: %s (%s)",
            SCENE_DISPLAY_NAMES.get(detection.scene_id, detection.scene_id),
            detection.reason,
        )
        return detection

    def _get_account_name(self) -> str:
        """
        从仪表盘页面提取当前登录的账号名称。
        通常在页面右上角头像旁边、或页面顶栏显示账号/店铺/达人名称。
        """
        candidates: list[str] = []

        def add_candidate(value: str):
            text = str(value or "").strip()
            if text:
                candidates.append(text)

        name_selectors = [
            # 右上角用户名/店铺名/达人名
            '[class*="user-name"]',
            '[class*="username"]',
            '[class*="shop-name"]',
            '[class*="shopName"]',
            '[class*="account-name"]',
            '[class*="accountName"]',
            '[class*="nick-name"]',
            '[class*="nickName"]',
            '[class*="talent-name"]',
            '[class*="talentName"]',
            # 头像旁的文字
            '[class*="avatar"] + span',
            '[class*="avatar"] + div',
            '[class*="header-user"] span',
            '[class*="header-right"] span',
            '[class*="user-info"] span',
            '[class*="userInfo"] span',
        ]
        for sel in name_selectors:
            try:
                el = self.page.query_selector(sel)
                if el and el.is_visible():
                    add_candidate(el.inner_text())
            except Exception:
                continue

        try:
            header_candidates = self.page.evaluate(
                r"""
                () => {
                  const normalize = (value) => (value || '').replace(/\s+/g, ' ').trim();
                  const viewportWidth = window.innerWidth || document.documentElement.clientWidth || 0;
                  const viewportHeight = window.innerHeight || document.documentElement.clientHeight || 0;
                  const nodes = Array.from(document.querySelectorAll('a, button, span, div'));
                  const values = [];
                  for (const node of nodes) {
                    const rect = node.getBoundingClientRect();
                    if (rect.width < 20 || rect.height < 16) continue;
                    if (rect.top < 0 || rect.top > Math.max(220, viewportHeight * 0.25)) continue;
                    if (rect.left < viewportWidth * 0.7) continue;
                    const text = normalize(node.innerText || '');
                    if (!text || text.length > 40) continue;
                    values.push(text);
                  }
                  return values.slice(0, 80);
                }
                """
            ) or []
            for value in header_candidates:
                add_candidate(value)
        except Exception:
            pass

        account_name = choose_account_name(candidates)
        if account_name:
            logger.info("检测到当前账号: %s", account_name)
            return account_name

        account_menu_selectors = [
            '[class*="avatar"]',
            '[class*="user-info"]',
            '[class*="userInfo"]',
            '[class*="account"]',
            '[class*="profile"]',
            '[class*="header-right"] [aria-haspopup]',
        ]
        for selector in account_menu_selectors:
            try:
                trigger = self.page.query_selector(selector)
                if trigger and trigger.is_visible():
                    trigger.click()
                    self.page.wait_for_timeout(500)
                    break
            except Exception:
                continue

        try:
            menu_candidates = self.page.evaluate(
                r"""
                () => {
                  const normalize = (value) => (value || '').replace(/\s+/g, ' ').trim();
                  const viewportWidth = window.innerWidth || document.documentElement.clientWidth || 0;
                  const nodes = Array.from(document.querySelectorAll('div, section, article'));
                  const panels = nodes
                    .filter((node) => {
                      const text = normalize(node.innerText || '');
                      if (!text.includes('退出登录') && !text.includes('切换子账号') && !text.includes('切换账号')) {
                        return false;
                      }
                      const rect = node.getBoundingClientRect();
                      return rect.width >= 180 && rect.height >= 120 && rect.left >= viewportWidth * 0.68;
                    })
                    .sort((left, right) => {
                      const l = left.getBoundingClientRect();
                      const r = right.getBoundingClientRect();
                      return (l.width * l.height) - (r.width * r.height);
                    });
                  if (!panels.length) return [];
                  return String(panels[0].innerText || '')
                    .split(/\n+/)
                    .map((line) => normalize(line))
                    .filter(Boolean);
                }
                """
            ) or []
            account_name = choose_account_name(menu_candidates)
            if account_name:
                logger.info("从右上角账号菜单检测到账号: %s", account_name)
                return account_name
        except Exception:
            pass

        # 备选：尝试从页面 title 中提取，但只作为最后兜底
        try:
            title = self.page.title()
            account_name = choose_account_name([], fallback_title=title or "")
            if account_name:
                logger.info("从页面标题检测到账号: %s", account_name)
                return account_name
        except Exception:
            pass

        logger.warning("未能自动识别当前账号名称")
        return ""

    # ============================================================
    # 导航 - 按场景进入目标页面
    # ============================================================
    def _navigate_to_target_scene(self):
        logger.info("准备进入目标场景: %s", SCENE_DISPLAY_NAMES.get(self.scene_id, self.scene_id))

        if self.scene_id == SCENE_HOME_OVERVIEW:
            self._navigate_home_overview()
            return

        if self.scene_id == SCENE_VIDEO_REVIEW:
            self._navigate_video_review()
            return

        if self.scene_id == SCENE_SHOP_LIVE_DATA:
            self._navigate_shop_live_data()
            return

        self._navigate_creator_live_review()

    def _navigate_home_overview(self):
        logger.info("导航到 首页数据...")
        selectors = [
            'nav a:has-text("首页")',
            'header a:has-text("首页")',
            '[class*="tab"]:has-text("首页")',
            '[class*="nav"] a:has-text("首页")',
            'a:has-text("首页")',
            'span:has-text("首页")',
        ]
        home_link = self._find_element(selectors, "首页菜单")
        home_link.click()
        wait_for_page_ready(self.page, timeout_ms=8000)
        logger.info("已进入首页页面")

    def _navigate_creator_live_review(self):
        """
        达人入口导航：绝对不打开新窗口。

        核心原理：
        "直播"链接是纯 JS onclick → window.open() 打开新窗口。
        在点击之前，拦截并替换 window.open，让它变成 window.location.href 跳转，
        这样就永远不会有新窗口出现。
        """
        logger.info("【达人端】导航到 直播 → 直播复盘（无新窗口）...")

        self.page.bring_to_front()
        self._scroll_to_top()
        self.page.wait_for_timeout(1500)

        # ----------------------------------------------------------------
        # 步骤 1：注入 JS，拦截 window.open，改为普通跳转
        # ----------------------------------------------------------------
        logger.info("步骤1：拦截 window.open，防止新窗口...")

        self.page.evaluate("""
            () => {
                const _origOpen = window.open;
                window.open = function(url) {
                    // 如果是直播相关的窗口，只做普通跳转，不开新窗口
                    if (url && url.includes('live')) {
                        window.location.href = url;
                        return null;  // 返回 null 表示没有真的打开窗口
                    }
                    return _origOpen.apply(this, arguments);
                };
            }
        """)
        logger.info("  window.open 已拦截")

        # ----------------------------------------------------------------
        # 步骤 2：点击"直播"链接
        # ----------------------------------------------------------------
        logger.info("步骤2：点击'直播'链接...")

        live_selectors = [
            'nav a:has-text("直播")',
            'header a:has-text("直播")',
            '[class*="tab"]:has-text("直播")',
            '[class*="nav"] a:has-text("直播")',
            '[class*="header"] a:has-text("直播")',
            '[class*="menu"] a:has-text("直播")',
            'a:has-text("直播")',
        ]

        try:
            live_nav = self._find_element(live_selectors, "直播菜单", max_attempts=3)
            live_nav.click()
            logger.info("已点击'直播'")
            wait_for_page_ready(self.page, timeout_ms=10000)
            self.page.wait_for_timeout(3000)
        except RuntimeError:
            logger.warning("未找到'直播'链接")
            self.page.wait_for_timeout(2000)

        # ----------------------------------------------------------------
        # 步骤 3：在当前窗口左侧菜单找"直播复盘"
        # ----------------------------------------------------------------
        logger.info("步骤3：在左侧菜单点击'直播复盘'...")

        review_selectors = [
            'aside [class*="menu"] a:has-text("直播复盘")',
            'aside [class*="nav"] a:has-text("直播复盘")',
            'aside a:has-text("直播复盘")',
            'aside [role="menu"] a:has-text("直播复盘")',
            'aside li:has-text("直播复盘")',
            '[class*="sidebar"] a:has-text("直播复盘")',
            '[class*="side-bar"] a:has-text("直播复盘")',
            '[class*="left-menu"] a:has-text("直播复盘")',
            'a:has-text("直播复盘")',
            'span:has-text("直播复盘")',
            'button:has-text("直播复盘")',
        ]

        review_clicked = False
        for sel in review_selectors:
            try:
                el = self._find_element([sel], f"直播复盘({sel})", max_attempts=2)
                if el:
                    box = el.bounding_box()
                    if box:
                        viewport_w = self.page.viewport_size()['width']
                        if box['x'] < viewport_w * 0.5:
                            href = el.get_attribute("href")
                            if href:
                                logger.info(f"  用 window.location.href 直接跳转: {href}")
                                self.page.evaluate(f"window.location.href = '{href}'")
                                wait_for_page_ready(self.page, timeout_ms=10000)
                                self.page.wait_for_timeout(3000)
                            else:
                                el.click()
                                wait_for_page_ready(self.page, timeout_ms=8000)
                                self.page.wait_for_timeout(2000)
                            review_clicked = True
                            logger.info(f"  已导航到'直播复盘': {sel}")
                            break
            except Exception:
                continue

        if not review_clicked:
            logger.warning("左侧未找到，全页尝试...")
            try:
                el = self._find_element(review_selectors, "直播复盘(全页)", max_attempts=3)
                if el:
                    href = el.get_attribute("href")
                    if href:
                        self.page.evaluate(f"window.location.href = '{href}'")
                        wait_for_page_ready(self.page, timeout_ms=10000)
                        self.page.wait_for_timeout(3000)
                    else:
                        el.click()
                        wait_for_page_ready(self.page, timeout_ms=8000)
                        self.page.wait_for_timeout(2000)
                    logger.info("已点击'直播复盘'（全页）")
                    review_clicked = True
            except RuntimeError:
                pass

        self._scroll_to_top()
        wait_for_page_ready(self.page, timeout_ms=10000)
        self.page.wait_for_timeout(3000)
        logger.info(f"  最终页面 URL: {self.page.url}")

        # ----------------------------------------------------------------
        # 步骤 4：确认没有多窗口
        # ----------------------------------------------------------------
        try:
            all_pages = self.browser.contexts[0].pages
            logger.info(f"导航完成后窗口数量: {len(all_pages)}（应为 1）")
        except Exception:
            pass

    def _navigate_video_review(self):
        """短视频入口导航：仅走顶部"短视频"，必要时直接访问视频复盘页面。"""
        logger.info("导航到 短视频 → 视频复盘...")
        self._scroll_to_top()
        self.page.wait_for_timeout(1000)

        video_selectors = [
            'nav a:has-text("短视频")',
            'header a:has-text("短视频")',
            '[class*="tab"]:has-text("短视频")',
            '[class*="nav"] a:has-text("短视频")',
            'a:has-text("短视频")',
            'span:has-text("短视频")',
        ]
        try:
            video_nav = self._find_element(video_selectors, "短视频菜单", max_attempts=3)
            video_nav.click()
            wait_for_page_ready(self.page, timeout_ms=8000)
            self._scroll_to_top()
        except RuntimeError:
            logger.info("未找到'短视频'标签，尝试直接访问...")

        self._scroll_to_top()
        self.page.wait_for_timeout(2000)

        # 多次检查页面上下文
        for attempt in range(3):
            if self._is_video_review_detail_page_context():
                logger.info(f'已通过顶部"短视频"进入视频复盘页面 (尝试 {attempt + 1})')
                return
            self.page.wait_for_timeout(2000)

        # 尝试直接访问视频复盘页面
        for target_url in self._build_video_review_entry_urls():
            logger.info("尝试直接访问视频复盘页面: %s", target_url)
            try:
                self.page.goto(target_url, wait_until="networkidle", timeout=15000)
                wait_for_page_ready(self.page, timeout_ms=8000)
                self._scroll_to_top()
                self.page.wait_for_timeout(2000)
                if self._is_video_review_detail_page_context():
                    logger.info("已直接进入视频复盘页面")
                    return
            except Exception as e:
                logger.warning(f"访问 {target_url} 失败: {e}")
                continue

        logger.warning('无法确认视频复盘页面上下文，继续尝试抓取...')

    def _navigate_shop_live_data(self):
        """
        店铺入口导航：顶部导航栏点击"直播" → 子菜单点击"实时直播数据"
        """
        logger.info("【店铺端】导航到 直播 → 实时直播数据...")

        # 第一步：点击顶部导航的"直播"标签
        live_selectors = [
            'nav a:has-text("直播")',
            'header a:has-text("直播")',
            '[class*="tab"]:has-text("直播")',
            '[class*="nav"] a:has-text("直播")',
            '[class*="header"] a:has-text("直播")',
            'a:has-text("直播")',
            'span:has-text("直播")',
        ]
        try:
            live_nav = self._find_element(live_selectors, "直播菜单")
            live_nav.click()
            wait_for_page_ready(self.page, timeout_ms=8000)
        except RuntimeError:
            logger.info("未找到'直播'标签，继续查找子菜单...")

        # 第二步：点击"实时直播数据"
        realtime_selectors = [
            'a:has-text("实时直播数据")',
            'span:has-text("实时直播数据")',
            '[class*="menu"]:has-text("实时直播数据")',
            'li:has-text("实时直播数据")',
        ]
        try:
            link = self._find_element(realtime_selectors, "实时直播数据")
            link.click()
            wait_for_page_ready(self.page, timeout_ms=8000)
            logger.info("已进入实时直播数据页面")
            return
        except RuntimeError:
            pass

        # 备选：直接通过 URL 访问
        url = self.config.get("compass_url", "https://compass.jinritemai.com")
        fallback_urls = [
            f"{url}/live/realtime",
            f"{url}/shop/live/realtime",
        ]
        for fb_url in fallback_urls:
            logger.info(f"尝试直接访问: {fb_url}")
            try:
                self.page.goto(fb_url, wait_until="networkidle", timeout=15000)
                wait_for_page_ready(self.page, timeout_ms=8000)
                if "404" not in self.page.title().lower():
                    logger.info(f"已通过 URL 进入实时直播数据页面")
                    return
            except Exception:
                continue

        raise RuntimeError(
            "无法进入实时直播数据页面。请确认店铺罗盘中是否有「直播 → 实时直播数据」菜单。"
        )

    # ============================================================

    def _scroll_page(self, steps: int = 6, distance: int = 800):
        for _ in range(steps):
            try:
                self.page.mouse.wheel(0, distance)
            except Exception:
                pass
            self.page.wait_for_timeout(800)

    def _scroll_to_top(self):
        try:
            self.page.evaluate("() => window.scrollTo(0, 0)")
            self.page.wait_for_timeout(400)
        except Exception:
            pass

    def _find_section_by_texts(self, texts: list[str], min_width: int = 900, min_height: int = 250):
        script = r"""
        (payload) => {
          const normalize = (value) => (value || '').replace(/\s+/g, '').trim();
          const targets = (payload.texts || []).map(normalize).filter(Boolean);
          const minWidth = payload.minWidth || 0;
          const minHeight = payload.minHeight || 0;
          const candidates = Array.from(document.querySelectorAll('section, div, article'));
          const matches = candidates.filter((node) => {
            const text = normalize(node.innerText);
            if (!text) return false;
            const rect = node.getBoundingClientRect();
            if (rect.width < minWidth || rect.height < minHeight) {
              return false;
            }
            return targets.every((target) => text.includes(target));
          });
          if (!matches.length) {
            return null;
          }
          matches.sort((a, b) => {
            const rectA = a.getBoundingClientRect();
            const rectB = b.getBoundingClientRect();
            const areaDiff = rectA.height * rectA.width - rectB.height * rectB.width;
            if (areaDiff !== 0) {
              return areaDiff;
            }
            return (a.innerText || '').length - (b.innerText || '').length;
          });
          return matches[0];
        }
        """
        try:
            handle = self.page.evaluate_handle(
                script,
                {"texts": texts, "minWidth": min_width, "minHeight": min_height},
            )
        except Exception:
            return None
        try:
            return handle.as_element()
        except Exception:
            return None

    def _get_scope_text(self, scope) -> str:
        try:
            return self.page.evaluate("(root) => root ? (root.innerText || '') : ''", scope) or ""
        except Exception:
            return ""

    def _ensure_tab_active(self, scope, text: str) -> None:
        selectors = [
            f'[role="tab"]:has-text("{text}")',
            f'button:has-text("{text}")',
            f'span:has-text("{text}")',
            f'a:has-text("{text}")',
        ]
        try:
            for sel in selectors:
                try:
                    el = scope.query_selector(sel)
                except Exception:
                    el = None
                if el and el.is_visible():
                    try:
                        el.click()
                        self.page.wait_for_timeout(1200)
                        wait_for_page_ready(self.page, timeout_ms=3000)
                    except Exception:
                        pass
                    return
        except Exception:
            return

    def _extract_metric_from_scope_text(
        self,
        scope_text: str,
        labels: tuple[str, ...],
        regex_patterns: tuple[str, ...] = (),
    ) -> tuple[float | None, str]:
        raw = extract_metric_fragment(scope_text, labels, regex_patterns) or ""
        return parse_metric_value(raw), raw

    def _extract_metrics_from_scope(self, scope, label_map: dict[str, str]) -> dict:
        script = r"""
        (root, labels) => {
          const normalize = (value) => (value || '').replace(/\s+/g, '').trim();
          const entries = {};
          const nodes = Array.from(root.querySelectorAll('div, li, section, article'));
          for (const node of nodes) {
            const raw = normalize(node.innerText);
            if (!raw) continue;
            for (const [label, key] of Object.entries(labels)) {
              if (!raw.includes(label)) continue;
              const lines = raw.split(/\n+/).map((line) => normalize(line)).filter(Boolean);
              const lineIndex = lines.findIndex((line) => line.includes(label));
              let valueCandidate = null;
              if (lineIndex >= 0) {
                valueCandidate = lines[lineIndex + 1] || lines[lineIndex + 2] || null;
              }
              if (!valueCandidate) {
                const match = raw.replace(label, "").match(/[-¥￥0-9,.]+[万亿]?/);
                valueCandidate = match ? match[0] : null;
              }
              if (valueCandidate && !entries[key]) {
                entries[key] = valueCandidate;
              }
            }
          }
          return entries;
        }
        """
        try:
            raw = self.page.evaluate(script, scope, label_map)
        except Exception:
            raw = {}

        parsed = {}
        for key, value in (raw or {}).items():
            parsed[key] = parse_metric_value(value)
            parsed[f"{key}_raw"] = value
        return parsed

    def _extract_table_rows(self, scope) -> list[list[str]]:
        script = r"""
        (root) => {
          const normalize = (value) => (value || '').replace(/\s+/g, ' ').trim();
          const rows = [];
          const pushRow = (cells) => {
            const cleaned = cells.map((cell) => normalize(cell)).filter(Boolean);
            if (cleaned.length >= 2) {
              rows.push(cleaned);
            }
          };

          const tables = Array.from(root.querySelectorAll('table'));
          for (const table of tables) {
            const trs = Array.from(table.querySelectorAll('tr'));
            for (const tr of trs) {
              const cells = Array.from(tr.querySelectorAll('th, td')).map((cell) => cell.innerText || '');
              pushRow(cells);
            }
          }

          if (rows.length) {
            return rows;
          }

          const roleRows = Array.from(root.querySelectorAll('[role="row"]'));
          for (const row of roleRows) {
            const cells = Array.from(row.querySelectorAll('[role="cell"], [role="gridcell"], td, th, div, span'))
              .map((cell) => cell.innerText || '');
            pushRow(cells);
          }

          if (rows.length) {
            return rows;
          }

          const listItems = Array.from(root.querySelectorAll('ul li'));
          for (const item of listItems) {
            const text = normalize(item.innerText || '');
            if (!text) {
              continue;
            }
            const parts = text.split(/\\s+/).filter(Boolean);
            if (parts.length >= 2) {
              rows.push(parts);
            }
          }

          return rows;
        }
        """
        try:
            rows = self.page.evaluate(script, scope)
        except Exception:
            rows = []
        return rows or []

    def _extract_home_overview_summary_metrics(self) -> dict:
        """
        抓取首页 → 整体概览卡片中的核心指标。
        提取：成交金额、成交订单数、退款金额（通过点击标签切换）。
        不提取：预估佣金收入。
        日期范围已在进入首页前通过 _apply_date_selection 设置。
        """
        metrics: dict[str, float | str | None] = {}

        # 滚动到顶部并等待数据加载
        self._scroll_to_top()
        self.page.wait_for_timeout(2000)

        # 定位整体概览卡片
        scope = self._find_home_overview_scope()

        # 定义要抓取的指标列表
        indicators = [
            ("成交金额", "成交金额"),
            ("成交订单数", "成交订单数"),
            ("退款金额", "退款金额"),
        ]

        for tab_name, metric_name in indicators:
            # 点击标签切换
            self._click_overview_tab(tab_name, scope)
            self.page.wait_for_timeout(2500)  # 等待数据加载

            # 点击后重新定位 scope
            scope = self._find_home_overview_scope()
            scope_text = self._get_scope_text(scope)

            # 提取数据
            value, raw = self._extract_metric_from_scope_text(
                scope_text, (metric_name,)
            )
            metrics[metric_name] = value
            metrics[f"{metric_name}_raw"] = raw
            logger.info("%s: %s", metric_name, value)

            # 在"成交金额" tab 下提前抓取板块分布（避免切到"退款金额"后读错数据）
            if tab_name == "成交金额":
                panel_metrics = self._extract_home_overview_panel_metrics()
                metrics.update(panel_metrics)

        logger.info(
            "首页整体概览提取完成: 成交金额=%s, 成交订单数=%s, 退款金额=%s",
            metrics.get("成交金额"), metrics.get("成交订单数"), metrics.get("退款金额"),
        )
        return metrics

    def _extract_home_overview_panel_metrics(self) -> dict:
        """
        抓取首页整体概况卡片中的板块分布表格。
        包含：直播成交金额、直播环比、短视频成交金额、短视频环比、商品卡成交金额、商品卡环比。
        """
        try:
            scope = self._find_home_overview_scope()
            text = self._get_scope_text(scope)
        except Exception:
            return {}

        metrics: dict[str, float | str | None] = {}

        # 解析板块分布表格：匹配 "直播  ¥xxx" 等行
        lines = text.split("\n")
        current_block = None
        for line in lines:
            line = line.strip()
            if not line:
                continue
            if line == "直播":
                current_block = "直播"
            elif line == "短视频":
                current_block = "短视频"
            elif line == "商品卡":
                current_block = "商品卡"
            elif current_block and "¥" in line:
                amount_match = re.search(r"¥\s*([\d,.]+)", line)
                if amount_match:
                    raw_amount = amount_match.group(0)
                    amount_val = float(amount_match.group(1).replace(",", ""))
                    metrics[f"{current_block}_成交金额"] = amount_val
                    metrics[f"{current_block}_成交金额_raw"] = raw_amount
                current_block = None

        for k in ("直播_成交金额", "短视频_成交金额", "商品卡_成交金额"):
            logger.info("%s: %s", k, metrics.get(k))
        return metrics

    def _extract_home_short_video_core_metrics(self) -> dict:
        self._scroll_page(steps=6, distance=900)
        scope = self._find_section_by_texts(["短视频", "视频复盘", "核心数据"])
        if not scope:
            raise RuntimeError("未找到首页短视频模块")

        metrics: dict[str, float | str | None] = {}

        self._ensure_tab_active(scope, "核心数据")
        core_text = self._get_scope_text(scope)

        leads_orders, leads_orders_raw = self._extract_metric_from_scope_text(
            core_text,
            ("引流成交次数", "引流成交订单数"),
            (r"共有([0-9.,]+(?:万|亿)?)人次通过短视频引流成功下单",),
        )
        metrics["short_video_leads_orders"] = leads_orders
        metrics["short_video_leads_orders_raw"] = leads_orders_raw

        leads_amount, leads_amount_raw = self._extract_metric_from_scope_text(
            core_text,
            ("引流成交金额",),
        )
        metrics["short_video_leads_amount"] = leads_amount
        metrics["short_video_leads_amount_raw"] = leads_amount_raw

        exposure_value = None
        exposure_raw = ""
        for tab_name in ("核心数据", "流量转化", "流量来源"):
            self._ensure_tab_active(scope, tab_name)
            tab_text = self._get_scope_text(scope)
            exposure_value, exposure_raw = self._extract_metric_from_scope_text(
                tab_text,
                ("曝光次数", "视频曝光次数", "曝光人数"),
            )
            if exposure_value is not None:
                break

        metrics["short_video_exposure"] = exposure_value
        metrics["short_video_exposure_raw"] = exposure_raw
        return metrics

    def _extract_video_review_page_metrics(self) -> dict:
        panel_scope = self._open_video_review_more_data_panel()
        tab_metrics = self._collect_video_review_summary_metrics(panel_scope)
        if not any((metrics or {}) for metrics in tab_metrics.values()):
            raise RuntimeError('未从短视频复盘"更多数据"弹层提取到任何指标卡片')
        return tab_metrics

    def _resolve_video_review_header_scope(self):
        self._assert_video_review_detail_page_context()
        scope = self._find_scope_by_text_groups(
            [
                (["视频表现", "更多数据", "近7天"], 1000, 220),
                (["视频表现", "更多数据"], 1000, 220),
                (["视频表现", "近1天", "近7天", "近30天"], 1000, 220),
            ]
        )
        if not scope:
            raise RuntimeError('未找到视频复盘页面顶部"视频表现"区域')
        return scope

    def _resolve_video_review_more_data_scope(self):
        self._assert_video_review_detail_page_context()
        scope = self._find_scope_by_text_groups(
            [
                (["全部数据", "直接成交", "引流价值", "数据趋势"], 1000, 700),
                (["全部数据", "直接成交", "引流价值", "引流成交金额"], 1000, 500),
                (["引流成交金额", "引流直播间成交金额", "引流直播间曝光次数"], 900, 400),
            ]
        )
        return scope

    def _assert_video_review_detail_page_context(self):
        if self.scene_id != SCENE_VIDEO_REVIEW:
            return

        if self._is_video_review_detail_page_context():
            return

        raise RuntimeError('当前页面不是"短视频 -> 视频复盘"独立页面，已终止抓取')

    def _is_video_review_detail_page_context(self) -> bool:
        visible_texts = self._visible_keywords_for_page(
            self.page,
            ("视频复盘", "视频明细", "视频榜单", "视频表现", "更多数据"),
        )
        return is_video_review_detail_page_snapshot(self.page.url, visible_texts)

    def _build_video_review_entry_urls(self) -> list[str]:
        raw_url = (self.page.url or self.config.get("compass_url") or "").strip()
        parts = urlsplit(raw_url)
        if parts.scheme and parts.netloc:
            origin = f"{parts.scheme}://{parts.netloc}"
        else:
            base = self.config.get("compass_url", "https://compass.jinritemai.com").strip()
            base_parts = urlsplit(base)
            origin = (
                f"{base_parts.scheme}://{base_parts.netloc}"
                if base_parts.scheme and base_parts.netloc
                else "https://compass.jinritemai.com"
            )
        return [
            f"{origin}/talent/video-analysis",
            f"{origin}/talent/video-analysis?from_page=%2Ftalent",
        ]

    def _open_video_review_more_data_panel(self):
        panel_scope = self._resolve_video_review_more_data_scope()
        panel_text = self._get_scope_text(panel_scope) if panel_scope else ""
        if panel_scope and "全部数据" in panel_text and "引流价值" in panel_text:
            return panel_scope

        header_scope = self._resolve_video_review_header_scope()
        selectors = [
            'button:has-text("更多数据")',
            'span:has-text("更多数据")',
            'div:has-text("更多数据")',
            'a:has-text("更多数据")',
        ]

        trigger = self._find_element(
            selectors,
            '短视频复盘"更多数据"按钮',
            target=header_scope,
            max_attempts=2,
            wait_ms=500,
        )

        if not trigger:
            raise RuntimeError('未找到短视频复盘"更多数据"按钮')

        trigger.click()
        self.page.wait_for_timeout(1200)
        wait_for_page_ready(self.page, timeout_ms=3000)

        panel_scope = self._resolve_video_review_more_data_scope()
        if not panel_scope:
            raise RuntimeError('未成功打开短视频复盘"更多数据"弹层')
        return panel_scope

    def _activate_video_review_more_data_tab(self, panel_scope, tab_name: str):
        selectors = [
            f'[role="tab"]:has-text("{tab_name}")',
            f'button:has-text("{tab_name}")',
            f'span:has-text("{tab_name}")',
            f'div:has-text("{tab_name}")',
            f'a:has-text("{tab_name}")',
        ]
        tab = self._find_element(
            selectors,
            f'短视频复盘"更多数据"弹层页签: {tab_name}',
            target=panel_scope,
            max_attempts=2,
            wait_ms=300,
        )
        tab.click()
        self.page.wait_for_timeout(800)
        wait_for_page_ready(self.page, timeout_ms=3000)

        if tab_name == "引流价值":
            wait_until(
                lambda: any(
                    marker in self._get_scope_text(panel_scope)
                    for marker in ("引流直播间曝光次数", "引流成交金额")
                ),
                timeout_seconds=5,
                interval_seconds=0.3,
                timeout_message='已打开"更多数据"弹层，但未成功切换到"引流价值"页签',
            )

    def _extract_all_metric_cards_from_scope(self, scope) -> dict[str, float | None]:
        script = r"""
        (root) => {
          const normalize = (value) => (value || '').replace(/\s+/g, ' ').trim();
          const excluded = new Set([
            '全部数据',
            '直接成交',
            '引流价值',
            '数据趋势',
            '流量来源分布',
            '下载明细',
            '取消',
          ]);
          const candidates = [];
          const nodes = Array.from(root.querySelectorAll('div, section, article, li'));

          for (const node of nodes) {
            const rect = node.getBoundingClientRect();
            if (rect.width < 120 || rect.height < 70 || rect.width > 420 || rect.height > 260) {
              continue;
            }
            const lines = String(node.innerText || '')
              .split(/\n+/)
              .map((line) => normalize(line))
              .filter(Boolean);
            if (lines.length < 2 || lines.length > 8) {
              continue;
            }

            const label = lines[0];
            if (!label || excluded.has(label)) {
              continue;
            }
            if (label.includes('较上周期') || label.includes('较同行')) {
              continue;
            }

            let value = '';
            for (const line of lines.slice(1)) {
              if (/[¥￥]?\d/.test(line)) {
                value = line;
                break;
              }
            }
            if (!value) {
              continue;
            }

            candidates.push({
              label,
              value,
              area: Math.round(rect.width * rect.height),
            });
          }

          candidates.sort((left, right) => left.area - right.area);
          const entries = {};
          for (const item of candidates) {
            if (!entries[item.label]) {
              entries[item.label] = item.value;
            }
          }
          return entries;
        }
        """
        try:
            raw = self.page.evaluate(script, scope) or {}
        except Exception:
            raw = {}

        metrics: dict[str, float | None] = {}
        for label, value in raw.items():
            metrics[str(label)] = parse_metric_value(str(value))
        return metrics

    def _collect_video_review_summary_metrics(self, panel_scope) -> dict[str, dict[str, float | None]]:
        tab_metrics: dict[str, dict[str, float | None]] = {}

        for tab_name in ("直接成交", "引流价值"):
            self._activate_video_review_more_data_tab(panel_scope, tab_name)
            metrics = self._extract_all_metric_cards_from_scope(panel_scope)
            if not metrics:
                raise RuntimeError(f'已打开"更多数据"弹层，但未读取到"{tab_name}"页签指标卡片')
            tab_metrics[tab_name] = metrics

        return tab_metrics

    def _extract_chart_options_from_scope(self, scope) -> list[dict]:
        try:
            charts = self.page.evaluate(
                """
                (root) => {
                  const echarts = window.echarts;
                  if (!echarts || typeof echarts.getInstanceByDom !== 'function') {
                    return [];
                  }

                  const seen = new Set();
                  const results = [];
                  const candidates = [root, ...Array.from(root.querySelectorAll('div, canvas'))];

                  for (const node of candidates) {
                    let current = node;
                    while (current) {
                      let instance = null;
                      try {
                        instance = echarts.getInstanceByDom(current);
                      } catch (error) {
                        instance = null;
                      }
                      if (instance) {
                        const ident = String(instance.id || results.length);
                        if (!seen.has(ident)) {
                          seen.add(ident);
                          const option = instance.getOption ? instance.getOption() : {};
                          const xAxis = Array.isArray(option.xAxis) ? option.xAxis : [option.xAxis || {}];
                          const series = Array.isArray(option.series) ? option.series : [];
                          results.push({
                            xAxisData: Array.isArray(xAxis[0]?.data) ? xAxis[0].data : [],
                            series: series.map((item) => ({
                              name: item?.name || '',
                              data: Array.isArray(item?.data)
                                ? item.data.map((point) => {
                                    if (point && typeof point === 'object' && 'value' in point) {
                                      return point.value;
                                    }
                                    return point;
                                  })
                                : [],
                            })),
                          });
                        }
                        break;
                      }
                      current = current.parentElement;
                    }
                  }

                  return results.sort((left, right) => right.xAxisData.length - left.xAxisData.length);
                }
                """,
                scope,
            )
        except Exception:
            return []
        return charts or []

    def _extract_live_review_page_metrics(self) -> dict:
        self._scroll_page(steps=2, distance=500)
        scope = self._find_scope_by_text_groups(
            [
                (["直播复盘", "问题诊断"], 1200, 450),
                (["直播复盘", "成交金额", "千次观看成交金额"], 1000, 260),
                (["直播", "直播复盘", "成交金额"], 1000, 260),
            ]
        )
        if not scope:
            scope = self.page.query_selector("body")
        if not scope:
            raise RuntimeError("未找到直播复盘模块")

        scope_text = self._get_scope_text(scope)
        metrics: dict[str, float | str | None] = {}

        parseable_metrics = {
            "成交金额": "live_review_summary_amount",
            "直播场次": "live_review_session_count",
            "开播天数": "live_review_live_days",
            "处罚次数": "live_review_penalty_count",
            "平台扶持流量": "live_review_supported_traffic",
            "单位小时曝光次数": "live_review_hourly_exposure",
            "曝光-观看率": "live_review_watch_rate",
            "千次观看成交金额": "live_review_thousand_view_amount",
        }

        for label, key in parseable_metrics.items():
            value, raw = self._extract_metric_from_scope_text(scope_text, (label,))
            metrics[key] = value
            metrics[f"{key}_raw"] = raw

        duration_raw = extract_metric_fragment(scope_text, ("直播时长",)) or ""
        metrics["live_review_duration_raw"] = duration_raw
        return metrics

    def _extract_live_review_channel_analysis_as_df(self) -> pd.DataFrame:
        """
        抓取直播复盘页面的渠道分析漏斗表格。

        表头（上→下/左→右）：
        渠道名称 / 直播间曝光次数 / 直播间观看次数 / 直播间曝光-观看率 /
        商品曝光次数 / 商品曝光-点击率 / 成交订单数 / 成交金额

        关键：页面有多个含"渠道"的卡片，要通过"直播间曝光次数"等关键词
        精确定位到漏斗数据表，而不是"成交体裁"那个表。
        """
        logger.info("[渠道分析] 开始抓取漏斗表格...")

        # 1. 滚动到页面顶部
        self._scroll_to_top()
        self.page.wait_for_timeout(1500)

        # 2. 找"渠道分析"四个字所在区域
        # 用"直播间曝光次数"做二次确认，排除"成交体裁"那个错误的表
        scope = self._find_scope_by_text_groups(
            [
                # 优先：同时含"渠道分析"和"直播间曝光次数"的区域
                (["渠道分析", "直播间曝光次数"], 1500, 400),
                # 其次：含"渠道分析"和"商品曝光"的区域
                (["渠道分析", "商品曝光"], 1500, 400),
                # 再次：含"渠道分析"和"直播间"的区域
                (["渠道分析", "直播间"], 1500, 400),
                # 兜底：只有"渠道分析"
                (["渠道分析"], 1200, 400),
            ]
        )
        if not scope:
            logger.warning("[渠道分析] 未找到渠道分析卡片区域")
            return pd.DataFrame()

        # 3. 滚动到卡片位置，触发懒加载
        try:
            box = scope.bounding_box()
            if box:
                viewport_h = self.page.viewport_size.get("height", 900) or 900
                target_y = box["y"] + box["height"] / 2 - viewport_h / 2
                target_y = max(0, target_y)
                logger.info(f"[渠道分析] 滚动到卡片位置: y={target_y:.0f}")
                self.page.evaluate(f"window.scrollTo(0, {target_y})")
                self.page.wait_for_timeout(2000)
        except Exception as e:
            logger.warning(f"[渠道分析] 滚动定位失败: {e}")

        # 4. 多次小步滚动，触发懒加载表格
        for _ in range(8):
            self.page.evaluate("window.scrollBy(0, 250)")
            self.page.wait_for_timeout(600)

        self.page.wait_for_timeout(2000)

        # 5. 再次定位，确保表格已加载
        scope = self._find_scope_by_text_groups(
            [
                (["渠道分析", "直播间曝光次数"], 1500, 400),
                (["渠道分析", "商品曝光"], 1500, 400),
                (["渠道分析", "直播间"], 1500, 400),
                (["渠道分析"], 1200, 400),
            ]
        )
        if not scope:
            logger.warning("[渠道分析] 再次查找仍未找到卡片")
            return pd.DataFrame()

        # 6. 提取表格
        rows = self._extract_table_rows(scope)
        if not rows:
            logger.warning("[渠道分析] 未提取到表格行，尝试获取原始文本")
            text = self._get_scope_text(scope)
            logger.info(f"[渠道分析] 原始文本: {text[:800]}")
            return pd.DataFrame()

        logger.info(f"[渠道分析] 提取到 {len(rows)} 行原始数据")
        logger.info(f"[渠道分析] 原始表头: {rows[0] if rows else []}")

        # 7. 解析行，构建 DataFrame
        if len(rows) < 2:
            logger.warning("[渠道分析] 数据行不足")
            return pd.DataFrame()

        header_row = rows[0]
        data_rows = rows[1:]

        # 标准化表头（直接映射，无模糊匹配，避免匹配到错误的表）
        # 表头列对应关系（按位置）：
        # 0:渠道名称 1:直播间曝光次数 2:直播间观看次数 3:直播间曝光-观看率
        # 4:商品曝光次数 5:商品曝光-点击率 6:成交订单数 7:成交金额
        std_header_map = {
            0: "渠道名称",
            1: "直播间曝光次数",
            2: "直播间观看次数",
            3: "直播间曝光-观看率",
            4: "商品曝光次数",
            5: "商品曝光-点击率",
            6: "成交订单数",
            7: "成交金额",
        }

        std_header = [std_header_map.get(i, h.strip()) for i, h in enumerate(header_row)]
        max_cols = len(std_header)

        logger.info(f"[渠道分析] 标准化表头: {std_header}")

        # 构建 DataFrame
        normalized_rows = []
        for row in data_rows:
            if len(row) >= max_cols:
                normalized_rows.append(row[:max_cols])
            else:
                normalized_rows.append(list(row) + [""] * (max_cols - len(row)))

        df = pd.DataFrame(normalized_rows, columns=std_header[:max_cols])
        df = df.dropna(how="all")

        # 过滤空行
        if "渠道名称" in df.columns:
            df = df[df["渠道名称"].notna() & (df["渠道名称"] != "")]
        else:
            df = df.dropna(how="all")

        return df

    def _extract_home_live_traffic_sources(self) -> dict:
        self._scroll_page(steps=6, distance=900)
        scope = self._find_section_by_texts(["直播", "直播复盘", "流量来源"])
        if not scope:
            raise RuntimeError("未找到首页直播模块")
        self._ensure_tab_active(scope, "流量来源")
        self.page.wait_for_timeout(1500)
        rows = self._extract_table_rows(scope)
        return {"rows": rows}

    def _export_home_overview_metrics(self) -> tuple[pd.DataFrame, Path]:
        logger.info("导出首页概览指标（直播流量来源 + 短视频核心数据）")
        metrics = {}
        rows = []
        try:
            rows = self._extract_home_live_traffic_sources().get("rows", [])
        except Exception as exc:
            logger.warning("直播流量来源提取失败: %s", exc)
        try:
            metrics.update(self._extract_home_short_video_core_metrics())
        except Exception as exc:
            logger.warning("短视频核心数据提取失败: %s", exc)

        if not metrics and not rows:
            raise RuntimeError("首页指标提取失败，未获取到任何数据")

        if rows:
            header = rows[0] if any("来源" in cell or "流量" in cell for cell in rows[0]) else None
            body = rows[1:] if header else rows
            max_len = max((len(row) for row in body), default=0)
            if header:
                columns = header + [f"指标{idx}" for idx in range(len(header) + 1, max_len + 1)]
            else:
                columns = [f"指标{idx}" for idx in range(1, max_len + 1)]
            padded = [row + [""] * (max_len - len(row)) for row in body]
            df = pd.DataFrame(padded, columns=columns)
        else:
            df = pd.DataFrame([{}])

        if "流量来源" in df.columns and "渠道名称" not in df.columns:
            df["渠道名称"] = df["流量来源"]
        if "成交金额" in df.columns and "用户支付金额" not in df.columns:
            df["用户支付金额"] = df["成交金额"]

        for key, value in metrics.items():
            df[key] = value

        df, csv_path = self._process_and_save_dataframe(df)
        logger.info("首页指标已保存: %s", csv_path)
        return df, csv_path

    def _process_and_save_dataframe(self, dataframe: pd.DataFrame) -> tuple[pd.DataFrame, Path]:
        portal_tag = "creator" if self.portal_type == self.PORTAL_CREATOR else "shop"
        return persist_exported_dataframe(
            dataframe,
            portal_type=portal_tag,
            account_name=self.account_name,
            config=self.config,
            task_metadata={
                "task_id": self.task_id,
                "scene_id": self.scene_id,
                "scene_name": SCENE_DISPLAY_NAMES.get(self.scene_id, self.scene_id),
                "date_mode": self.target_date_range.mode if self.target_date_range else "",
                "target_start_date": (
                    self.target_date_range.start.isoformat() if self.target_date_range else ""
                ),
                "target_end_date": (
                    self.target_date_range.end.isoformat() if self.target_date_range else ""
                ),
            },
        )

    # 数据导出和保存
    # ============================================================
    def _apply_date_selection(self):
        """按场景和日期策略选择页面日期。

        - 近一天：点击"自定义" → 选择前一天（如 2026/04/14-2026/04/14）
        - 近七天：点击"近七天"快捷按钮
        覆盖所有场景（包括首页整体概括）。
        """
        if not self.target_date_range:
            raise RuntimeError("任务日期范围未初始化")

        self.date_scope = self._resolve_scene_date_scope()

        logger.info(
            "开始设置日期: 场景=%s, 模式=%s, 目标=%s ~ %s",
            SCENE_DISPLAY_NAMES.get(self.scene_id, self.scene_id),
            self.target_date_range.label,
            self.target_date_range.start.isoformat(),
            self.target_date_range.end.isoformat(),
        )

        if self.target_date_range.mode == DATE_MODE_LAST_1_DAY:
            # 近一天：点击"自定义" → 选择单日
            self._open_single_day_picker_mode(required=True, target=self.date_scope)
            self._select_single_day_from_picker(self.target_date_range.start, target=self.date_scope)
            return

        # 近七天：点击快捷按钮
        self._select_quick_date_range(
            "近七天",
            ["近7天", "最近7天", "近七天", "最近七天", "近7日", "最近7日"],
            target=self.date_scope,
        )

    def _find_home_overview_scope(self):
        try:
            handle = self.page.evaluate_handle(
                """
                () => {
                  const candidates = Array.from(document.querySelectorAll('div, section, article'))
                    .filter((node) => {
                      const text = (node.innerText || '').replace(/\\s+/g, ' ').trim();
                      if (!(text.includes('整体概况') || text.includes('整体概览')) || !text.includes('更多')) {
                        return false;
                      }
                      const rect = node.getBoundingClientRect();
                      return rect.width >= 280 && rect.height >= 120;
                    })
                    .sort((left, right) => {
                      const leftArea = left.getBoundingClientRect().width * left.getBoundingClientRect().height;
                      const rightArea = right.getBoundingClientRect().width * right.getBoundingClientRect().height;
                      return leftArea - rightArea;
                    });
                  return candidates[0] || null;
                }
                """
            )
            element = handle.as_element()
            if element:
                logger.info('已定位首页"整体概况"卡片')
                return element
        except Exception:
            pass

        raise RuntimeError('未定位到首页"整体概况"卡片，无法安全执行首页数据抓取')

    def _find_scope_by_text_groups(
        self,
        groups: list[tuple[list[str], int, int]],
    ):
        for texts, min_width, min_height in groups:
            scope = self._find_section_by_texts(
                texts,
                min_width=min_width,
                min_height=min_height,
            )
            if scope:
                return scope
        return None

    def _resolve_scene_date_scope(self):
        scope_candidates: list[tuple[list[str], int, int]] = []

        if self.scene_id == SCENE_VIDEO_REVIEW:
            self._assert_video_review_detail_page_context()
            scope_candidates = [
                (["视频表现", "更多数据", "近1天", "近7天", "近30天"], 1000, 220),
                (["视频表现", "近1天", "近7天", "近30天"], 1000, 220),
            ]
        elif self.scene_id == SCENE_LIVE_REVIEW:
            scope_candidates = [
                (["直播复盘", "近7天", "自定义"], 1000, 220),
                (["直播复盘", "问题诊断"], 1200, 500),
                (["直播", "直播复盘", "成交金额"], 1000, 260),
            ]
        elif self.scene_id == SCENE_HOME_OVERVIEW:
            try:
                return self._find_home_overview_scope()
            except Exception:
                return None

        return self._find_scope_by_text_groups(scope_candidates)

    def _open_home_overview_more_filters(self, scope):
        more_selectors = [
            'button:has-text("更多")',
            'span:has-text("更多")',
            'div:has-text("更多")',
            'a:has-text("更多")',
        ]

        try:
            trigger = self._find_element(
                more_selectors,
                '首页"整体概况"卡片内的"更多"按钮',
                target=scope,
                max_attempts=2,
            )
            # 使用 JavaScript 点击避免元素脱离 DOM 的问题
            self.page.evaluate("(el) => el.click()", trigger)
            wait_for_page_ready(self.page, timeout_ms=3000)
            logger.info('已展开首页"整体概况"卡片的"更多"筛选区')
        except Exception as e:
            logger.warning(f"点击'更多'按钮失败: {e}，继续提取数据")

    def _click_overview_tab(self, tab_name: str, scope):
        """
        点击整体概览卡片中的指标标签（成交金额/成交订单数/退款金额）
        标签不在 scope 内，需要在整页搜索
        """
        logger.info(f"尝试点击标签: {tab_name}")

        # 标签选择器 - 按优先级排序
        tab_selectors = [
            f'[role="tab"]:has-text("{tab_name}")',
            f'[role="button"]:has-text("{tab_name}")',
            f'button:has-text("{tab_name}")',
            f'span:has-text("{tab_name}")',
            f'div:has-text("{tab_name}")',
            f'a:has-text("{tab_name}")',
        ]

        for attempt in range(3):
            try:
                # 在整页查找标签
                trigger = self._find_element(
                    tab_selectors,
                    f'整体概览"{tab_name}"标签',
                    max_attempts=2,
                )
                if trigger:
                    # 使用 JavaScript 点击
                    self.page.evaluate("(el) => el.click()", trigger)
                    wait_for_page_ready(self.page, timeout_ms=2500)
                    logger.info(f"已点击\"{tab_name}\"标签")
                    return
            except Exception as e:
                logger.warning(f"点击\"{tab_name}\"标签第{attempt + 1}次失败: {e}")
                self.page.wait_for_timeout(500)

        logger.warning(f"点击\"{tab_name}\"标签失败，继续执行")

    def _open_single_day_picker_mode(self, *, required: bool, target=None):
        selectors = [
            'button:has-text("自定义")',
            'span:has-text("自定义")',
            'div:has-text("自定义")',
            'a:has-text("自定义")',
        ]

        if target is not None:
            try:
                trigger = self._find_element(
                    selectors,
                    "当前场景单日日期入口",
                    target=target,
                    max_attempts=6,
                    wait_ms=300,
                )
                # 使用 JavaScript 点击避免打开新窗口
                self.page.evaluate("(el) => el.click()", trigger)
                wait_for_page_ready(self.page, timeout_ms=3000)
                logger.info("已在当前场景切换到单日日期选择面板")
                return
            except Exception:
                pass

        for selector in selectors:
            try:
                trigger = self.page.query_selector(selector)
                if trigger and trigger.is_visible():
                    # 使用 JavaScript 点击避免打开新窗口
                    self.page.evaluate("(el) => el.click()", trigger)
                    wait_for_page_ready(self.page, timeout_ms=3000)
                    logger.info("已切换到单日日期选择面板")
                    return
            except Exception:
                continue

        if required:
            raise RuntimeError("当前页面未找到单日日期入口，无法安全设置昨天数据")

    def _select_quick_date_range(self, display_name: str, texts: list[str], target=None):
        logger.info(f"选择日期范围: {display_name}")
        selectors = []
        for text in texts:
            selectors.extend([
                f'span:has-text("{text}")',
                f'button:has-text("{text}")',
                f'div:has-text("{text}")',
                f'li:has-text("{text}")',
                f'a:has-text("{text}")',
                f'[role="tab"]:has-text("{text}")',
                f'[role="button"]:has-text("{text}")',
                f'text="{text}"',
            ])

        try:
            btn = self._find_element(
                selectors,
                f"{display_name}按钮",
                target=target,
                max_attempts=20,
                wait_ms=500,
            )
        except RuntimeError as exc:
            logger.info("快捷日期选择器未命中，尝试无障碍/文本回退: %s", exc)
            btn = self._find_element_by_role(texts, target=target)
            if not btn:
                btn = self._find_element_by_text(texts, target=target)
            if not btn:
                raise

        try:
            if hasattr(btn, "scroll_into_view_if_needed"):
                btn.scroll_into_view_if_needed()
        except Exception:
            pass

        btn.click()
        wait_for_page_ready(self.page, timeout_ms=5000)

    def _select_single_day_from_picker(self, target_date: date, target=None):
        """尽量通过页面日期控件将当前场景切换到指定单日。"""
        date_str = target_date.strftime("%Y-%m-%d")
        logger.info("选择单日日期: %s", date_str)

        trigger_selectors = [
            'input[placeholder*="日期"]',
            'input[placeholder*="开始"]',
            'input[placeholder*="选择"]',
            '[class*="date-picker"]',
            '[class*="picker"] input',
            '[class*="range"] input',
        ]

        search_roots = [target, self.page] if target is not None else [self.page]
        trigger_opened = False

        for root in search_roots:
            if root is None:
                continue
            for selector in trigger_selectors:
                try:
                    el = root.query_selector(selector)
                    if el and el.is_visible():
                        # 使用 JavaScript 点击
                        self.page.evaluate("(elem) => elem.click()", el)
                        wait_for_page_ready(self.page, timeout_ms=3000)
                        trigger_opened = True
                        break
                except Exception:
                    continue
            if trigger_opened:
                break

        if not trigger_opened:
            for selector in trigger_selectors:
                try:
                    el = self.page.query_selector(selector)
                    if el and el.is_visible():
                        # 使用 JavaScript 点击
                        self.page.evaluate("(elem) => elem.click()", el)
                        wait_for_page_ready(self.page, timeout_ms=3000)
                        break
                except Exception:
                    continue

        range_inputs = self.page.locator(
            'input[placeholder*="开始"], '
            'input[placeholder*="结束"], '
            'input[placeholder*="日期"], '
            'input[placeholder*="选择日期"], '
            'input[placeholder*="Start"], '
            'input[placeholder*="End"]'
        )

        try:
            count = range_inputs.count()
        except Exception:
            count = 0

        if count >= 2:
            first = range_inputs.nth(0)
            second = range_inputs.nth(1)
            self._fill_date_input(first, date_str)
            self._fill_date_input(second, date_str)
            second.press("Enter")
            wait_for_page_ready(self.page, timeout_ms=5000)
            return

        if count == 1:
            only = range_inputs.nth(0)
            if self._try_fill_single_date_input(only, date_str):
                return

        raise RuntimeError(f"无法自动设置页面日期，请手动选择 {date_str} 后再执行")

    def _fill_date_input(self, locator, value: str):
        locator.click()
        locator.fill("")
        locator.fill(value)

    def _try_fill_single_date_input(self, locator, date_str: str) -> bool:
        attempts = [date_str, f"{date_str} - {date_str}"]
        for attempt in attempts:
            try:
                locator.click()
                locator.fill("")
                locator.fill(attempt)
                locator.press("Enter")
                wait_for_page_ready(self.page, timeout_ms=5000)
                return True
            except Exception:
                continue

        return False

    def _collect_date_haystacks(self, scope=None) -> list[str]:
        if scope is not None:
            try:
                haystacks = self.page.evaluate(
                    """
                    (root) => {
                      const values = [];
                      for (const input of Array.from(root.querySelectorAll('input')).slice(0, 8)) {
                        const value = (input.value || '').trim();
                        if (value) {
                          values.push(value);
                        }
                      }
                      const rootText = (root.innerText || '').trim();
                      if (rootText) {
                        values.push(rootText);
                      }
                      return values;
                    }
                    """,
                    scope,
                )
                if haystacks:
                    return haystacks
            except Exception:
                pass

        haystacks = []
        try:
            inputs = self.page.locator("input")
            count = inputs.count()
            for idx in range(min(count, 8)):
                loc = inputs.nth(idx)
                try:
                    if loc.is_visible(timeout=500):
                        value = loc.input_value().strip()
                        if value:
                            haystacks.append(value)
                    else:
                        continue
                except Exception:
                    continue
        except Exception:
            pass

        try:
            date_text_blocks = self.page.locator(
                '[class*="date"], [class*="picker"], [class*="range"], [class*="calendar"]'
            )
            count = date_text_blocks.count()
            for idx in range(min(count, 8)):
                loc = date_text_blocks.nth(idx)
                try:
                    if loc.is_visible(timeout=500):
                        text = loc.inner_text().strip()
                        if text:
                            haystacks.append(text)
                except Exception:
                    continue
        except Exception:
            pass

        try:
            page_text = self.page.evaluate(
                """
                () => {
                  const root = document.querySelector('main') || document.body || document.documentElement;
                  return root ? (root.innerText || '').trim() : '';
                }
                """
            )
            if page_text:
                haystacks.append(page_text)
        except Exception:
            pass

        return haystacks

    def _assert_date_selection_applied(self, target_range):
        if self.date_scope is not None:
            haystacks = self._collect_date_haystacks(scope=self.date_scope)
            if is_target_date_range_visible(target_range.start, target_range.end, haystacks):
                logger.info(
                    "页面日期已确认切换到目标范围: %s ~ %s（整体概况卡片）",
                    target_range.start.isoformat(),
                    target_range.end.isoformat(),
                )
                return target_range
            if self.scene_id in {SCENE_LIVE_REVIEW, SCENE_VIDEO_REVIEW, SCENE_SHOP_LIVE_DATA}:
                preview = " | ".join(haystacks[:3]) if haystacks else "未读到当前场景日期文本"
                logger.warning(
                    "当前场景日期作用域未直接匹配目标范围 %s ~ %s，继续回退整页检查。当前检测到: %s",
                    target_range.start.isoformat(),
                    target_range.end.isoformat(),
                    preview,
                )

        haystacks = self._collect_date_haystacks()

        if is_target_date_range_visible(target_range.start, target_range.end, haystacks):
            logger.info(
                "页面日期已确认切换到目标范围: %s ~ %s",
                target_range.start.isoformat(),
                target_range.end.isoformat(),
            )
            return target_range

        if self.scene_id == SCENE_HOME_OVERVIEW:
            logger.warning("首页日期范围未匹配目标值，继续使用当前页面范围")
            return target_range

        raise RuntimeError(
            "页面上未检测到目标日期范围 "
            f"{target_range.start.isoformat()} ~ {target_range.end.isoformat()}，"
            "为避免导出错误数据，本次任务已终止"
        )

    def _export_data(self) -> Path:
        logger.info("导出数据...")
        selectors = [
            'button:has-text("导出")', 'span:has-text("导出")',
            'a:has-text("导出")', '[class*="export"]',
            'button:has-text("下载")',
        ]
        export_btn = self._find_element(selectors, "导出按钮")

        timeout_ms = self.config.get("download_timeout", 120) * 1000
        with self.page.expect_download(timeout=timeout_ms) as dl_info:
            export_btn.click()

        download = dl_info.value
        original_name = download.suggested_filename or "直播数据.xlsx"
        date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        stem = Path(original_name).stem
        suffix = Path(original_name).suffix or ".xlsx"
        save_path = DOWNLOAD_DIR / f"{stem}_{date_str}{suffix}"

        download.save_as(str(save_path))
        logger.info(f"文件已下载: {save_path}")
        return save_path

    def _export_video_review_metrics(self) -> tuple[pd.DataFrame, Path]:
        if not self.target_date_range:
            raise RuntimeError("短视频复盘缺少目标日期范围")

        logger.info('导出短视频复盘"更多数据"弹层中的全部指标卡片')
        summary_metrics = self._extract_video_review_page_metrics()
        rows = build_video_review_export_rows(self.target_date_range, summary_metrics)
        df = pd.DataFrame(rows)
        df, csv_path = self._process_and_save_dataframe(df)
        logger.info("短视频复盘页面指标已保存: %s", csv_path)
        return df, csv_path

    def _apply_video_review_date_input(self):
        """
        短视频引流专用日期设置：直接操作 input[placeholder="开始日期"] /
        input[placeholder="结束日期"] 输入框，填入目标日期后点击确定。
        兼容近一天和近七天模式。
        """
        if not self.target_date_range:
            raise RuntimeError("短视频引流缺少目标日期范围")

        date_start = self.target_date_range.start.strftime("%Y-%m-%d")
        date_end = self.target_date_range.end.strftime("%Y-%m-%d")

        logger.info(f"短视频日期输入: {date_start} ~ {date_end}")

        try:
            start_input = self.page.locator('input[placeholder="开始日期"]').first
            end_input = self.page.locator('input[placeholder="结束日期"]').first

            # 先点击"近一天"快捷按钮
            self._select_quick_date_range(
                "近一天",
                ["近1天", "最近1天", "近一天", "最近一天", "近1日", "最近1日"],
                target=None,
            )
            wait_for_page_ready(self.page, timeout_ms=3000)
            logger.info("短视频：已尝试选择近一天")

        except Exception:
            pass

        # 直接填充日期输入框
        try:
            start_input = self.page.locator('input[placeholder="开始日期"]').first
            end_input = self.page.locator('input[placeholder="结束日期"]').first
            start_input.click(timeout=5000)
            start_input.fill(date_start)
            end_input.click(timeout=5000)
            end_input.fill(date_end)

            # 点击确定
            for sel in ['button:has-text("确定")', 'span:has-text("确定")',
                        '[class*="confirm"]', 'button:has-text("确认")']:
                try:
                    self.page.locator(sel).last.click(timeout=3000)
                    break
                except Exception:
                    continue

            wait_for_page_ready(self.page, timeout_ms=5000)
            logger.info(f"短视频日期设置完成: {date_start} ~ {date_end}")
        except Exception as exc:
            logger.warning("短视频日期输入失败（继续尝试抓取）: %s", exc)

    def _build_live_overall_summary(self, summary_metrics: dict) -> pd.DataFrame:
        """
        根据首页整体概括抓取的指标，构建"直播整体"Sheet 的 DataFrame。
        每账号一行数据，包含：账号名称、目标日期、抓取时间、成交金额、成交订单数、退款金额、
        直播成交金额、直播环比、短视频成交金额、短视频环比、商品卡成交金额、商品卡环比。
        """
        row = {"账号名称": self.account_name or ""}
        if self.target_date_range:
            row["目标日期"] = self.target_date_range.start.isoformat()
        else:
            row["目标日期"] = ""
        row["抓取时间"] = datetime.now().isoformat()

        # 成交金额
        row["成交金额"] = summary_metrics.get("成交金额")
        # 成交订单数
        row["成交订单数"] = summary_metrics.get("成交订单数")
        # 退款金额
        row["退款金额"] = summary_metrics.get("退款金额")

        # 板块分布（直播/短视频/商品卡 × 成交金额）
        row["直播成交金额"] = summary_metrics.get("直播_成交金额")
        row["短视频成交金额"] = summary_metrics.get("短视频_成交金额")
        row["商品卡成交金额"] = summary_metrics.get("商品卡_成交金额")

        df = pd.DataFrame([row])
        return df

    def _build_live_review_export_data(self) -> tuple[pd.DataFrame, Path]:
        """
        渠道明细：从直播复盘页面下载 Excel，读取并处理后返回 DataFrame。
        包含日期过滤、空列删除、元数据列添加。
        """
        filepath = self._export_data()
        df = pd.read_excel(filepath, engine="openpyxl")
        logger.info("渠道明细读取 %s 行, %s 列", len(df), len(df.columns))

        # 近一天模式按日期过滤
        if (
            self.target_date_range
            and self.target_date_range.mode == DATE_MODE_LAST_1_DAY
            and "日期" in df.columns
        ):
            target_iso = self.target_date_range.start.isoformat()
            normalized = df["日期"].map(normalize_export_row_date)
            filtered = df.loc[normalized == target_iso].copy()
            if not filtered.empty:
                logger.info("渠道明细近一天过滤: %s -> %s 行", len(df), len(filtered))
                df = filtered
            else:
                logger.warning("未匹配到目标日期，保留原始数据")

        # 删除所有空列（值全为 NaN / 空字符串 / 0 / None）
        empty_cols = []
        for col in df.columns:
            vals = df[col].dropna()
            vals = vals[vals.astype(str).str.strip() != ""]
            vals = vals[vals != 0]
            if len(vals) == 0:
                empty_cols.append(col)
        if empty_cols:
            df = df.drop(columns=empty_cols)
            logger.info("  已移除空列: %s", empty_cols)

        # 添加元数据列
        crawl_time = datetime.now().isoformat()
        if "账号名称" not in df.columns:
            df.insert(0, "账号名称", self.account_name or "")
        if "抓取时间" not in df.columns:
            df.insert(1, "抓取时间", crawl_time)
        if "目标日期" not in df.columns:
            df.insert(2, "目标日期", (
                self.target_date_range.start.isoformat()
                if self.target_date_range else ""
            ))

        return df, filepath

    def _write_excel_unified(self, xlsx_path: Path, sheets_data: list[dict]):
        """
        将多个 DataFrame 写入同一个 Excel 文件，每个 DataFrame 一个 Sheet。
        用于达人入口统一抓取后的合并导出。
        """
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        logger.info(f"写入多 Sheet Excel: {xlsx_path}")

        # 定义样式
        header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        alt_row_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
        data_alignment = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            for sheet_info in sheets_data:
                name = sheet_info.get("name", "Sheet")
                df = sheet_info.get("df")
                if df is None or (isinstance(df, pd.DataFrame) and df.empty):
                    logger.warning(f"Sheet '{name}' 无数据，跳过")
                    continue

                df.to_excel(writer, sheet_name=name, index=False)
                logger.info(f"  已写入 Sheet '{name}': {len(df)} 行")

                # 格式化已写入的工作表
                ws = writer.sheets[name]

                # 设置表头样式
                for col_idx, cell in enumerate(ws[1], 1):
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border

                # 设置数据行样式（交替颜色）
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
                    fill = alt_row_fill if row_idx % 2 == 0 else None
                    for cell in row:
                        cell.alignment = data_alignment
                        cell.border = thin_border
                        if fill:
                            cell.fill = fill

                # 自动调整列宽
                for col_idx, column_cells in enumerate(ws.columns, 1):
                    max_length = 0
                    column_letter = get_column_letter(col_idx)
                    for cell in column_cells:
                        try:
                            cell_len = len(str(cell.value)) if cell.value else 0
                            max_length = max(max_length, cell_len)
                        except Exception:
                            pass
                    adjusted_width = min(max_length + 4, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

                # 设置行高
                ws.row_dimensions[1].height = 25
                for row_idx in range(2, ws.max_row + 1):
                    ws.row_dimensions[row_idx].height = 20

    def _close(self):
        """关闭浏览器上下文，并在需要时终止外部 Chrome 进程。"""
        if self.context:
            try:
                self.context.close()
            except Exception:
                pass
            self.context = None

        if self.browser:
            try:
                self.browser.close()
            except Exception:
                pass
            self.browser = None

        if self.chrome_process:
            try:
                self.chrome_process.terminate()
                self.chrome_process.wait(timeout=5)
            except Exception:
                try:
                    self.chrome_process.kill()
                except Exception:
                    pass
            self.chrome_process = None

        logger.info("浏览器已关闭")

    def _iter_search_targets(self, root):
        """遍历页面及其子 Frame，用于查找元素。"""
        seen = set()
        queue = [root]

        while queue:
            current = queue.pop(0)
            if not current:
                continue
            ident = id(current)
            if ident in seen:
                continue
            seen.add(ident)
            yield current

            frames = []
            for attr in ("frames", "child_frames"):
                if not hasattr(current, attr):
                    continue
                try:
                    value = getattr(current, attr)
                    value = value() if callable(value) else value
                except Exception:
                    value = None
                if value:
                    frames = list(value)
                    break

            for frame in frames:
                if frame is current:
                    continue
                queue.append(frame)


    def _find_element_by_role(self, texts: list[str], target=None):
        target = target or self.page
        for search_target in self._iter_search_targets(target):
            if not hasattr(search_target, "get_by_role"):
                continue
            for text in texts:
                for role in ("tab", "button", "link"):
                    try:
                        loc = search_target.get_by_role(role, name=text)
                        if loc and loc.first.is_visible(timeout=1200):
                            return loc.first
                    except Exception:
                        continue
        return None

    def _find_element_by_text(self, texts: list[str], target=None):
        target = target or self.page
        script = r"""
        (candidates) => {
          const normalize = (value) => (value || '').replace(/\s+/g, '').trim();
          const wanted = candidates.map(normalize).filter(Boolean);
          const selectors = ['button', '[role="tab"]', '[role="button"]', 'a', 'li', 'span', 'div'];
          const elements = Array.from(document.querySelectorAll(selectors.join(',')));
          const isVisible = (el) => {
            const style = window.getComputedStyle(el);
            if (!style || style.display === 'none' || style.visibility === 'hidden' || style.opacity === '0') {
              return false;
            }
            const rect = el.getBoundingClientRect();
            return rect.width > 0 && rect.height > 0;
          };

          for (const wantedText of wanted) {
            const exact = elements.find((el) => isVisible(el) && normalize(el.innerText) === wantedText);
            if (exact) {
              return exact;
            }
          }

          for (const wantedText of wanted) {
            const partial = elements.find((el) => isVisible(el) && normalize(el.innerText).includes(wantedText));
            if (partial) {
              return partial;
            }
          }

          return null;
        }
        """
        for search_target in self._iter_search_targets(target):
            if not hasattr(search_target, "evaluate_handle"):
                continue
            try:
                handle = search_target.evaluate_handle(script, texts)
            except Exception:
                continue
            try:
                element = handle.as_element()
            except Exception:
                element = None
            if element:
                return element
        return None

    def _find_element(self, selectors: list, name: str, target=None, *, max_attempts: int = 1, wait_ms: int = 0):
        target = target or self.page
        attempts = max(1, max_attempts)

        for attempt in range(attempts):
            search_targets = list(self._iter_search_targets(target))
            for search_target in search_targets:
                for sel in selectors:
                    try:
                        el = search_target.query_selector(sel)
                        if el and el.is_visible():
                            return el
                    except Exception:
                        continue

            for search_target in search_targets:
                if not hasattr(search_target, "locator"):
                    continue
                for sel in selectors:
                    try:
                        loc = search_target.locator(sel).first
                        if loc.is_visible(timeout=3000):
                            return loc
                    except Exception:
                        continue

            if attempt < attempts - 1 and wait_ms > 0:
                try:
                    if self.page:
                        self.page.wait_for_timeout(wait_ms)
                    else:
                        time.sleep(wait_ms / 1000)
                except Exception:
                    time.sleep(wait_ms / 1000)

        raise RuntimeError(f"无法找到: {name}")
